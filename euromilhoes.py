#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          EUROMILHÕES - GERADOR PROFISSIONAL DE CHAVES v9.0                  ║
║          Baseado na metodologia Lotterycodex BI-BP-AI-AP                    ║
╚══════════════════════════════════════════════════════════════════════════════╝

Estratégias implementadas:
  - 16 Padrões Equilibrados (Lotterycodex BI/BP/AI/AP)
  - Filtro A: Soma configurável (padrão 80–190 | apertado 95–160)
  - Filtro B: Máx 1 par consecutivo, 0 triplos
  - Filtro C: Máx 2 números com mesmo dígito final
  - Filtro D: Mín 3 décadas diferentes
  - Filtro E: Máx 2 repetições do sorteio anterior
  - Filtro F: Sistema de cores (últimos 9 sorteios)
  - Filtro G: Regra do 31 – mín 1 número > 31 (anti-calendário)
  - Filtro H: Rejeitar progressões aritméticas perfeitas
  - Estrelas Equilibradas: sempre as 2 menos usadas
"""

import os
import sys
import json
import random
import sqlite3
import datetime
import csv
import time
from collections import Counter
from pathlib import Path

# ─── Dependency check ────────────────────────────────────────────────────────
MISSING_DEPS = []
try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    MISSING_DEPS.append("requests beautifulsoup4")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING_DEPS.append("openpyxl")

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text
    from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn
    from rich.prompt import Prompt, IntPrompt, Confirm
    from rich.layout import Layout
    from rich.columns import Columns
    from rich import box
    from rich.rule import Rule
    from rich.style import Style
except ImportError:
    MISSING_DEPS.append("rich")

if MISSING_DEPS and not os.environ.get("VERCEL"):
    print("\n[ERRO] Dependências em falta. Instala com:")
    print(f"  pip install {' '.join(MISSING_DEPS)}\n")
    sys.exit(1)

# ─── Date correction (Euromilhões draws only on Tuesdays and Fridays) ─────────
def corrigir_data_sorteio(data_str: str) -> str:
    """If date is not a Tuesday or Friday, move back to the nearest previous Tue/Fri."""
    d = datetime.date.fromisoformat(data_str)
    weekday = d.weekday()  # Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    if weekday in (1, 4):
        return data_str
    corrections = {0: 3, 2: 1, 3: 2, 5: 1, 6: 2}
    d -= datetime.timedelta(days=corrections[weekday])
    return d.isoformat()


# ─── EuroMillions Prize Tiers ────────────────────────────────────────────────
# (matched_numbers, matched_stars) → tier info
# Tiers 1-7 are pari-mutuel (variable); tiers 8-13 have typical fixed values.
PRIZE_TIERS = {
    (5, 2): {"tier": 1,  "name": "5+2 Jackpot",  "default_eur": None},
    (5, 1): {"tier": 2,  "name": "5+1",           "default_eur": None},
    (5, 0): {"tier": 3,  "name": "5+0",           "default_eur": None},
    (4, 2): {"tier": 4,  "name": "4+2",           "default_eur": None},
    (4, 1): {"tier": 5,  "name": "4+1",           "default_eur": None},
    (3, 2): {"tier": 6,  "name": "3+2",           "default_eur": None},
    (4, 0): {"tier": 7,  "name": "4+0",           "default_eur": None},
    (2, 2): {"tier": 8,  "name": "2+2",           "default_eur": 17.08},
    (3, 1): {"tier": 9,  "name": "3+1",           "default_eur": 12.73},
    (3, 0): {"tier": 10, "name": "3+0",           "default_eur": 10.58},
    (1, 2): {"tier": 11, "name": "1+2",           "default_eur": 9.52},
    (2, 1): {"tier": 12, "name": "2+1",           "default_eur": 7.19},
    (2, 0): {"tier": 13, "name": "2+0",           "default_eur": 4.28},
}

CUSTO_POR_APOSTA = 2.50  # € per single combination

# ─── Configurable constants (replace magic numbers) ──────────────────────────
COLOR_WINDOW_DRAWS = 9         # Number of recent draws for the colour system (Filter F)
MAX_HISTORICO_ENTRIES = 20     # Max entries kept in generation history
MAX_FAVORITES = 50             # Max saved favourite keys
MAX_SCRAPE_BATCH = 50          # Max dates per bulk prize-scrape run
SCRAPE_DELAY = 1.5             # Seconds between HTTP requests (rate-limit)
REQUEST_TIMEOUT = 8 if os.environ.get("VERCEL") else 20

# ─── User-Agent rotation pool ────────────────────────────────────────────────
_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
]


def _random_user_agent() -> str:
    """Return a randomly chosen User-Agent string."""
    return random.choice(_USER_AGENTS)


def _retry_request(url: str, headers: dict = None, timeout: int = None,
                   max_retries: int = 3, **kwargs) -> "requests.Response":
    """GET *url* with exponential-backoff retries.

    Raises the last exception if all attempts fail.
    """
    timeout = timeout or REQUEST_TIMEOUT
    if headers is None:
        headers = {}
    # Rotate User-Agent on each call
    headers = {**headers, "User-Agent": _random_user_agent()}

    last_exc: Exception | None = None
    for attempt in range(max_retries):
        try:
            return requests.get(url, headers=headers, timeout=timeout, **kwargs)
        except requests.exceptions.Timeout as exc:
            last_exc = exc
        except requests.exceptions.ConnectionError as exc:
            last_exc = exc
        if attempt < max_retries - 1:
            time.sleep(2 ** attempt)          # 1s, 2s between retries
    raise last_exc  # type: ignore[misc]


def validate_draw_numbers(nums, stars):
    """Validate draw numbers and stars. Returns (ok, error_msg)."""
    if not isinstance(nums, (list, tuple)) or len(nums) != 5:
        return False, "Precisa de exactamente 5 números"
    if not all(isinstance(n, int) and 1 <= n <= 50 for n in nums):
        return False, "Números devem ser inteiros entre 1 e 50"
    if len(set(nums)) != 5:
        return False, "Números não podem repetir"
    if not isinstance(stars, (list, tuple)) or len(stars) != 2:
        return False, "Precisa de exactamente 2 estrelas"
    if not all(isinstance(s, int) and 1 <= s <= 12 for s in stars):
        return False, "Estrelas devem ser inteiros entre 1 e 12"
    if len(set(stars)) != 2:
        return False, "Estrelas não podem repetir"
    return True, ""


# ─── Global constants ─────────────────────────────────────────────────────────
VERSION = "9.0"
_IS_VERCEL = bool(os.environ.get("VERCEL"))

if _IS_VERCEL:
    DB_PATH   = Path("/tmp/euromilhoes.db")
    EXCEL_DIR = Path("/tmp/chaves_geradas")
else:
    DB_PATH   = Path(__file__).parent / "euromilhoes.db"
    EXCEL_DIR = Path(__file__).parent / "chaves_geradas"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

console = Console()

# Lotterycodex quadrants
BI = [n for n in range(1, 26) if n % 2 != 0]   # Low-Odd:  1,3,5..25
BP = [n for n in range(2, 25) if n % 2 == 0]    # Low-Even: 2,4,6..24
AI = [n for n in range(27, 50) if n % 2 != 0]   # High-Odd: 27,29..49
AP = [n for n in range(26, 51) if n % 2 == 0]   # High-Even:26,28..50

# 16 balanced patterns [BI, BP, AI, AP]
PADROES_EQUILIBRADOS = [
    [1,1,1,2],[2,1,1,1],[1,2,1,1],[1,1,2,1],
    [2,2,0,1],[1,0,2,2],[2,0,1,2],[2,1,0,2],
    [2,0,2,1],[1,2,0,2],[2,2,1,0],[0,2,1,2],
    [0,1,2,2],[2,1,2,0],[0,2,2,1],[1,2,2,0],
]

# ── Full BI-BP-AI-AP pattern table (56 patterns, Lotterycodex) ────────────────
# Each entry: num, padrao [BI,BP,AI,AP], combinacoes, pct, grupo
# grupos: "best" | "middle_high" | "middle_low" | "worst"
TODOS_PADROES_BIBPAIAP = [
    # ── BEST (4) ──────────────────────────────────────────────────────────────
    {"num":  1, "padrao": [1,1,1,2], "combinacoes": 146014, "pct": 6.89, "grupo": "best"},
    {"num":  2, "padrao": [2,1,1,1], "combinacoes": 146016, "pct": 6.89, "grupo": "best"},
    {"num":  3, "padrao": [1,2,1,1], "combinacoes": 133848, "pct": 6.32, "grupo": "best"},
    {"num":  4, "padrao": [1,1,2,1], "combinacoes": 133848, "pct": 6.32, "grupo": "best"},
    # ── MIDDLE-HIGH (12) ──────────────────────────────────────────────────────
    {"num":  5, "padrao": [2,0,1,2], "combinacoes":  73008, "pct": 3.45, "grupo": "middle_high"},
    {"num":  6, "padrao": [2,1,0,2], "combinacoes":  73008, "pct": 3.45, "grupo": "middle_high"},
    {"num":  7, "padrao": [2,2,0,1], "combinacoes":  66924, "pct": 3.16, "grupo": "middle_high"},
    {"num":  8, "padrao": [1,0,2,2], "combinacoes":  66924, "pct": 3.16, "grupo": "middle_high"},
    {"num":  9, "padrao": [2,0,2,1], "combinacoes":  66924, "pct": 3.16, "grupo": "middle_high"},
    {"num": 10, "padrao": [1,2,0,2], "combinacoes":  66924, "pct": 3.16, "grupo": "middle_high"},
    {"num": 11, "padrao": [2,2,1,0], "combinacoes":  61776, "pct": 2.92, "grupo": "middle_high"},
    {"num": 12, "padrao": [0,2,1,2], "combinacoes":  61776, "pct": 2.92, "grupo": "middle_high"},
    {"num": 13, "padrao": [0,1,2,2], "combinacoes":  61776, "pct": 2.92, "grupo": "middle_high"},
    {"num": 14, "padrao": [2,1,2,0], "combinacoes":  61776, "pct": 2.92, "grupo": "middle_high"},
    {"num": 15, "padrao": [0,2,2,1], "combinacoes":  56628, "pct": 2.67, "grupo": "middle_high"},
    {"num": 16, "padrao": [1,2,2,0], "combinacoes":  56628, "pct": 2.67, "grupo": "middle_high"},
    # ── MIDDLE-LOW (12) ───────────────────────────────────────────────────────
    {"num": 17, "padrao": [3,0,1,1], "combinacoes":  44616, "pct": 2.11, "grupo": "middle_low"},
    {"num": 18, "padrao": [3,1,0,1], "combinacoes":  44616, "pct": 2.11, "grupo": "middle_low"},
    {"num": 19, "padrao": [1,0,1,3], "combinacoes":  44616, "pct": 2.11, "grupo": "middle_low"},
    {"num": 20, "padrao": [1,1,0,3], "combinacoes":  44616, "pct": 2.11, "grupo": "middle_low"},
    {"num": 21, "padrao": [3,1,1,0], "combinacoes":  41184, "pct": 1.94, "grupo": "middle_low"},
    {"num": 22, "padrao": [0,1,1,3], "combinacoes":  41184, "pct": 1.94, "grupo": "middle_low"},
    {"num": 23, "padrao": [1,0,3,1], "combinacoes":  37180, "pct": 1.75, "grupo": "middle_low"},
    {"num": 24, "padrao": [1,3,0,1], "combinacoes":  37180, "pct": 1.75, "grupo": "middle_low"},
    {"num": 25, "padrao": [0,3,1,1], "combinacoes":  34320, "pct": 1.62, "grupo": "middle_low"},
    {"num": 26, "padrao": [0,1,3,1], "combinacoes":  34320, "pct": 1.62, "grupo": "middle_low"},
    {"num": 27, "padrao": [1,1,3,0], "combinacoes":  34320, "pct": 1.62, "grupo": "middle_low"},
    {"num": 28, "padrao": [1,3,1,0], "combinacoes":  34320, "pct": 1.62, "grupo": "middle_low"},
    # ── WORST (28) ────────────────────────────────────────────────────────────
    {"num": 29, "padrao": [2,0,0,3], "combinacoes":  22308, "pct": 1.05, "grupo": "worst"},
    {"num": 30, "padrao": [3,0,0,2], "combinacoes":  22308, "pct": 1.05, "grupo": "worst"},
    {"num": 31, "padrao": [0,0,2,3], "combinacoes":  18876, "pct": 0.89, "grupo": "worst"},
    {"num": 32, "padrao": [3,2,0,0], "combinacoes":  18876, "pct": 0.89, "grupo": "worst"},
    {"num": 33, "padrao": [0,2,0,3], "combinacoes":  18876, "pct": 0.89, "grupo": "worst"},
    {"num": 34, "padrao": [3,0,2,0], "combinacoes":  18876, "pct": 0.89, "grupo": "worst"},
    {"num": 35, "padrao": [0,0,3,2], "combinacoes":  17160, "pct": 0.81, "grupo": "worst"},
    {"num": 36, "padrao": [0,3,0,2], "combinacoes":  17160, "pct": 0.81, "grupo": "worst"},
    {"num": 37, "padrao": [2,3,0,0], "combinacoes":  17160, "pct": 0.81, "grupo": "worst"},
    {"num": 38, "padrao": [2,0,3,0], "combinacoes":  17160, "pct": 0.81, "grupo": "worst"},
    {"num": 39, "padrao": [0,3,2,0], "combinacoes":  14520, "pct": 0.69, "grupo": "worst"},
    {"num": 40, "padrao": [0,2,3,0], "combinacoes":  14520, "pct": 0.69, "grupo": "worst"},
    {"num": 41, "padrao": [1,0,0,4], "combinacoes":   9296, "pct": 0.44, "grupo": "worst"},
    {"num": 42, "padrao": [4,0,0,1], "combinacoes":   9296, "pct": 0.44, "grupo": "worst"},
    {"num": 43, "padrao": [4,1,0,0], "combinacoes":   8580, "pct": 0.40, "grupo": "worst"},
    {"num": 44, "padrao": [0,0,1,4], "combinacoes":   8580, "pct": 0.40, "grupo": "worst"},
    {"num": 45, "padrao": [4,0,1,0], "combinacoes":   8580, "pct": 0.40, "grupo": "worst"},
    {"num": 46, "padrao": [0,1,0,4], "combinacoes":   8580, "pct": 0.40, "grupo": "worst"},
    {"num": 47, "padrao": [0,4,0,1], "combinacoes":   6435, "pct": 0.30, "grupo": "worst"},
    {"num": 48, "padrao": [1,4,0,0], "combinacoes":   6435, "pct": 0.30, "grupo": "worst"},
    {"num": 49, "padrao": [1,0,4,0], "combinacoes":   6435, "pct": 0.30, "grupo": "worst"},
    {"num": 50, "padrao": [0,0,4,1], "combinacoes":   6435, "pct": 0.30, "grupo": "worst"},
    {"num": 51, "padrao": [0,4,1,0], "combinacoes":   5940, "pct": 0.28, "grupo": "worst"},
    {"num": 52, "padrao": [0,1,4,0], "combinacoes":   5940, "pct": 0.28, "grupo": "worst"},
    {"num": 53, "padrao": [5,0,0,0], "combinacoes":   1287, "pct": 0.06, "grupo": "worst"},
    {"num": 54, "padrao": [0,0,0,5], "combinacoes":   1287, "pct": 0.06, "grupo": "worst"},
    {"num": 55, "padrao": [0,0,5,0], "combinacoes":    792, "pct": 0.04, "grupo": "worst"},
    {"num": 56, "padrao": [0,5,0,0], "combinacoes":    792, "pct": 0.04, "grupo": "worst"},
]

# Lookup dict for fast classification: tuple(padrao) → pattern entry
_PADRAO_LOOKUP = {tuple(p["padrao"]): p for p in TODOS_PADROES_BIBPAIAP}

_BI_SET = set(BI)
_BP_SET = set(BP)
_AI_SET = set(AI)
_AP_SET = set(AP)


def classificar_padrao_bibpaiap(numeros: list) -> dict | None:
    """Return the TODOS_PADROES_BIBPAIAP entry for a 5-number combination."""
    key = (
        sum(1 for n in numeros if n in _BI_SET),
        sum(1 for n in numeros if n in _BP_SET),
        sum(1 for n in numeros if n in _AI_SET),
        sum(1 for n in numeros if n in _AP_SET),
    )
    return _PADRAO_LOOKUP.get(key)


def classificar_padrao_cores(numeros: list, cores: dict) -> str:
    """Return e.g. '2210' representing [V, v, A, C] counts for the drawn numbers."""
    v = sum(1 for n in numeros if n in cores.get("vermelhos", set()))
    g = sum(1 for n in numeros if n in cores.get("verdes", set()))
    a = sum(1 for n in numeros if n in cores.get("azuis", set()))
    c = sum(1 for n in numeros if n in cores.get("castanhos", set()))
    return f"{v}{g}{a}{c}"


# Colour system thresholds
# VERMELHOS: 0 appearances in last 9 draws → 1–3 in combo
# VERDES: 1 appearance → 1–3 in combo
# AZUIS: 2 appearances → 0–2 in combo
# CASTANHOS: 3+ appearances → EXCLUDED (0 in combo)

# ═════════════════════════════════════════════════════════════════════════════
# DATABASE MANAGER
# ═════════════════════════════════════════════════════════════════════════════
class DatabaseManager:
    def __init__(self, db_path: Path = DB_PATH):
        self.db_path = db_path
        self._init_db()

    def _init_db(self):
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS sorteios (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    data        TEXT NOT NULL UNIQUE,
                    n1 INTEGER NOT NULL, n2 INTEGER NOT NULL, n3 INTEGER NOT NULL,
                    n4 INTEGER NOT NULL, n5 INTEGER NOT NULL,
                    e1 INTEGER NOT NULL, e2 INTEGER NOT NULL,
                    soma        INTEGER,
                    fonte       TEXT DEFAULT 'manual'
                );
                CREATE TABLE IF NOT EXISTS metadata (
                    chave TEXT PRIMARY KEY,
                    valor TEXT
                );
                CREATE TABLE IF NOT EXISTS chaves_geradas (
                    id      INTEGER PRIMARY KEY AUTOINCREMENT,
                    data_geracao TEXT,
                    numeros TEXT,
                    estrelas TEXT,
                    soma    INTEGER,
                    padrao  TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_sorteios_data ON sorteios(data);
                CREATE INDEX IF NOT EXISTS idx_sorteios_soma ON sorteios(soma);
                CREATE INDEX IF NOT EXISTS idx_metadata_chave ON metadata(chave);

                CREATE TABLE IF NOT EXISTS premios (
                    data        TEXT PRIMARY KEY
                                REFERENCES sorteios(data) ON DELETE CASCADE,
                    jackpot     REAL,
                    t1_prize REAL, t1_winners INTEGER,
                    t2_prize REAL, t2_winners INTEGER,
                    t3_prize REAL, t3_winners INTEGER,
                    t4_prize REAL, t4_winners INTEGER,
                    t5_prize REAL, t5_winners INTEGER,
                    t6_prize REAL, t6_winners INTEGER,
                    t7_prize REAL, t7_winners INTEGER,
                    t8_prize REAL, t8_winners INTEGER,
                    t9_prize REAL, t9_winners INTEGER,
                    t10_prize REAL, t10_winners INTEGER,
                    t11_prize REAL, t11_winners INTEGER,
                    t12_prize REAL, t12_winners INTEGER,
                    t13_prize REAL, t13_winners INTEGER
                );
            """)

    def eliminar_sorteio(self, data: str) -> bool:
        """Delete a draw by date. Returns True if a row was deleted."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute("DELETE FROM sorteios WHERE data = ?", (data,))
                return cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"[DB] eliminar_sorteio erro: {e}")
            return False

    def inserir_sorteio(self, data: str, numeros: list, estrelas: list, fonte: str = "manual") -> bool:
        """Insert a draw. Returns True if newly inserted, False if already existed or on error."""
        data = corrigir_data_sorteio(data)
        nums = sorted(numeros)
        ests = sorted(estrelas)
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute(
                    "INSERT OR IGNORE INTO sorteios (data,n1,n2,n3,n4,n5,e1,e2,soma,fonte) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (data, nums[0], nums[1], nums[2], nums[3], nums[4],
                     ests[0], ests[1], sum(nums), fonte)
                )
                return cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"[DB] inserir_sorteio erro: {e}")
            return False

    def ultimo_sorteio(self) -> dict | None:
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(
                "SELECT data,n1,n2,n3,n4,n5,e1,e2,soma FROM sorteios ORDER BY data DESC LIMIT 1"
            ).fetchone()
        if row:
            return {"data": row[0], "numeros": [row[1],row[2],row[3],row[4],row[5]],
                    "estrelas": [row[6],row[7]], "soma": row[8]}
        return None

    def ultimos_n_sorteios(self, n: int = COLOR_WINDOW_DRAWS) -> list[dict]:
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute(
                "SELECT data,n1,n2,n3,n4,n5,e1,e2 FROM sorteios ORDER BY data DESC LIMIT ?", (n,)
            ).fetchall()
        result = []
        for row in rows:
            result.append({"data": row[0], "numeros": [row[1],row[2],row[3],row[4],row[5]],
                            "estrelas": [row[6],row[7]]})
        return result

    def todos_sorteios(self) -> list[dict]:
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute(
                "SELECT data,n1,n2,n3,n4,n5,e1,e2,soma FROM sorteios ORDER BY data DESC"
            ).fetchall()
        result = []
        for row in rows:
            result.append({"data": row[0], "numeros": [row[1],row[2],row[3],row[4],row[5]],
                            "estrelas": [row[6],row[7]], "soma": row[8]})
        return result

    def total_sorteios(self) -> int:
        with sqlite3.connect(self.db_path) as conn:
            return conn.execute("SELECT COUNT(*) FROM sorteios").fetchone()[0]

    def guardar_chave_gerada(self, numeros: list, estrelas: list, padrao: list):
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                "INSERT INTO chaves_geradas (data_geracao,numeros,estrelas,soma,padrao) VALUES (?,?,?,?,?)",
                (datetime.datetime.now().isoformat(),
                 json.dumps(sorted(numeros)), json.dumps(sorted(estrelas)),
                 sum(numeros), json.dumps(padrao))
            )

    def get_metadata(self, chave: str) -> str | None:
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute("SELECT valor FROM metadata WHERE chave=?", (chave,)).fetchone()
        return row[0] if row else None

    def set_metadata(self, chave: str, valor: str):
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("INSERT OR REPLACE INTO metadata (chave,valor) VALUES (?,?)", (chave, valor))

    # ── Prize data methods ──────────────────────────────────────────────────
    def inserir_premios(self, data: str, prize_data: dict) -> bool:
        cols = ["data", "jackpot"]
        vals = [data, prize_data.get("jackpot")]
        for t in range(1, 14):
            cols += [f"t{t}_prize", f"t{t}_winners"]
            vals += [prize_data.get(f"t{t}_prize"), prize_data.get(f"t{t}_winners")]
        placeholders = ",".join(["?"] * len(vals))
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute(
                    f"INSERT OR REPLACE INTO premios ({','.join(cols)}) VALUES ({placeholders})",
                    vals
                )
            return True
        except sqlite3.Error as e:
            print(f"[DB] inserir_premios erro: {e}")
            return False

    def obter_premios(self, data: str) -> dict | None:
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM premios WHERE data = ?", (data,)).fetchone()
        if not row:
            return None
        return dict(row)

    def todos_premios(self) -> list[dict]:
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM premios ORDER BY data DESC").fetchall()
        return [dict(r) for r in rows]

    def tem_premios(self, data: str) -> bool:
        with sqlite3.connect(self.db_path) as conn:
            return conn.execute(
                "SELECT 1 FROM premios WHERE data = ?", (data,)
            ).fetchone() is not None

    def datas_sem_premios(self) -> list[str]:
        with sqlite3.connect(self.db_path) as conn:
            rows = conn.execute(
                "SELECT s.data FROM sorteios s LEFT JOIN premios p ON s.data = p.data "
                "WHERE p.data IS NULL ORDER BY s.data DESC"
            ).fetchall()
        return [r[0] for r in rows]


# ═════════════════════════════════════════════════════════════════════════════
# PRIZE CHECKER
# ═════════════════════════════════════════════════════════════════════════════
class PrizeChecker:
    """Verifies single or multiple-combination bets against a draw result."""

    MAX_NUMEROS = 15
    MAX_ESTRELAS = 12

    @staticmethod
    def verificar_aposta(
        numeros_jogados: list[int],
        estrelas_jogadas: list[int],
        sorteio_numeros: list[int],
        sorteio_estrelas: list[int],
        premios: dict | None = None,
    ) -> dict:
        from itertools import combinations

        n_nums = len(numeros_jogados)
        n_stars = len(estrelas_jogadas)

        if n_nums < 5 or n_nums > PrizeChecker.MAX_NUMEROS:
            return {"erro": f"Números devem ser entre 5 e {PrizeChecker.MAX_NUMEROS}"}
        if n_stars < 2 or n_stars > PrizeChecker.MAX_ESTRELAS:
            return {"erro": f"Estrelas devem ser entre 2 e {PrizeChecker.MAX_ESTRELAS}"}
        if not all(1 <= n <= 50 for n in numeros_jogados):
            return {"erro": "Números devem estar entre 1 e 50"}
        if not all(1 <= e <= 12 for e in estrelas_jogadas):
            return {"erro": "Estrelas devem estar entre 1 e 12"}
        if len(set(numeros_jogados)) != n_nums:
            return {"erro": "Números duplicados não são permitidos"}
        if len(set(estrelas_jogadas)) != n_stars:
            return {"erro": "Estrelas duplicadas não são permitidas"}

        sorteio_nums_set = set(sorteio_numeros)
        sorteio_stars_set = set(sorteio_estrelas)

        combos_nums = list(combinations(numeros_jogados, 5))
        combos_stars = list(combinations(estrelas_jogadas, 2))
        total_combinacoes = len(combos_nums) * len(combos_stars)
        custo_total = total_combinacoes * CUSTO_POR_APOSTA

        # Count wins by tier
        resultados_tier = {}
        for tier_key, tier_info in PRIZE_TIERS.items():
            resultados_tier[tier_info["tier"]] = {
                "tier": tier_info["tier"],
                "nome": tier_info["name"],
                "match": f"{tier_key[0]}+{tier_key[1]}",
                "quantidade": 0,
                "premio_unitario": 0.0,
                "subtotal": 0.0,
            }

        for combo_n in combos_nums:
            acertos_n = len(set(combo_n) & sorteio_nums_set)
            for combo_s in combos_stars:
                acertos_s = len(set(combo_s) & sorteio_stars_set)
                tier_info = PRIZE_TIERS.get((acertos_n, acertos_s))
                if tier_info:
                    t = tier_info["tier"]
                    resultados_tier[t]["quantidade"] += 1

        # Assign prize values from DB or defaults
        ganhos_totais = 0.0
        detalhe = []
        premios_em_falta = False
        for t in range(1, 14):
            rt = resultados_tier[t]
            if rt["quantidade"] == 0:
                continue
            # Get prize value: from DB data first, then default
            premio = 0.0
            premio_disponivel = True
            if premios and premios.get(f"t{t}_prize") is not None:
                val = premios[f"t{t}_prize"]
                # Treat 0.0 for tiers 1-7 as missing (pari-mutuel prizes can't be 0)
                if t <= 7 and (val == 0 or val == 0.0):
                    premio_disponivel = False
                    premios_em_falta = True
                else:
                    premio = float(val)
            else:
                # Find default from PRIZE_TIERS
                found_default = False
                for key, info in PRIZE_TIERS.items():
                    if info["tier"] == t and info["default_eur"] is not None:
                        premio = info["default_eur"]
                        found_default = True
                        break
                if not found_default:
                    premio_disponivel = False
                    premios_em_falta = True

            rt["premio_unitario"] = premio
            rt["premio_disponivel"] = premio_disponivel
            rt["subtotal"] = premio * rt["quantidade"]
            ganhos_totais += rt["subtotal"]
            detalhe.append(rt)

        return {
            "total_combinacoes": total_combinacoes,
            "custo_total": round(custo_total, 2),
            "ganhos_totais": round(ganhos_totais, 2),
            "lucro_prejuizo": round(ganhos_totais - custo_total, 2),
            "detalhe": detalhe,
            "premios_em_falta": premios_em_falta,
            "numeros_jogados": sorted(numeros_jogados),
            "estrelas_jogadas": sorted(estrelas_jogadas),
            "sorteio_numeros": sorted(sorteio_numeros),
            "sorteio_estrelas": sorted(sorteio_estrelas),
        }


# ═════════════════════════════════════════════════════════════════════════════
# PRIZE SCRAPER
# ═════════════════════════════════════════════════════════════════════════════
class PremiosScraper:
    """Scrapes prize breakdown data from euro-millions.com"""

    HEADERS = {
        "Accept-Language": "en-GB,en;q=0.5",
    }

    # Map tier names from the website to our tier numbers
    # Supports both "5 + 2" and "5+2" formats
    TIER_MAP = {
        "5 + 2": 1, "5+2": 1,
        "5 + 1": 2, "5+1": 2,
        "5": 3, "5 + 0": 3, "5+0": 3,
        "4 + 2": 4, "4+2": 4,
        "4 + 1": 5, "4+1": 5,
        "3 + 2": 6, "3+2": 6,
        "4": 7, "4 + 0": 7, "4+0": 7,
        "2 + 2": 8, "2+2": 8,
        "3 + 1": 9, "3+1": 9,
        "3": 10, "3 + 0": 10, "3+0": 10,
        "1 + 2": 11, "1+2": 11,
        "2 + 1": 12, "2+1": 12,
        "2": 13, "2 + 0": 13, "2+0": 13,
    }

    def scrape_premios(self, data_iso: str) -> dict | None:
        """Scrape prize breakdown for a single draw date (YYYY-MM-DD format)."""
        try:
            d = datetime.date.fromisoformat(data_iso)
            url = f"https://www.euro-millions.com/results/{d.strftime('%d-%m-%Y')}"
            resp = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
            if resp.status_code != 200:
                return None
            return self._parse_prizes(resp.text)
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            return None
        except Exception:
            return None

    def _parse_prizes(self, html: str) -> dict | None:
        """Parse prize breakdown table from HTML.
        Prefers the Portuguese table (€ values); falls back to first € table."""
        import re
        soup = BeautifulSoup(html, "html.parser")

        # Find the Portuguese prize table (has "Portuguese Winners" header)
        table = None
        for t in soup.find_all("table"):
            header_text = t.get_text().lower()
            if "portuguese" in header_text and "winners" in header_text:
                table = t
                break

        # Fallback: first table with € amounts and match/prize columns
        if not table:
            for t in soup.find_all("table"):
                text = t.get_text()
                if ("match" in text.lower() or "+" in text) and "€" in text:
                    table = t
                    break

        # Last fallback: any table with match + winners
        if not table:
            for t in soup.find_all("table"):
                text = t.get_text().lower()
                if "match" in text and ("prize" in text or "winners" in text):
                    table = t
                    break

        if not table:
            return None

        result = {}
        # Determine column layout from header row
        header_row = table.find("tr")
        headers = [h.get_text(strip=True).lower() for h in header_row.find_all(["td", "th"])] if header_row else []
        # "Total Winners" is the last column — use it for global winners count
        total_winners_col = None
        for idx, h in enumerate(headers):
            if "total" in h and "winner" in h:
                total_winners_col = idx

        rows = table.find_all("tr")
        for row in rows:
            cells = row.find_all(["td", "th"])
            if len(cells) < 3:
                continue

            match_text = cells[0].get_text(strip=True)
            # Normalize: "Match 5 + 2" → "5+2", strip whitespace around +
            match_text = re.sub(r"(?i)match\s*", "", match_text).strip()

            tier = self.TIER_MAP.get(match_text)
            if tier is None:
                continue

            # Prize Per Winner is column 1
            prize_text = cells[1].get_text(strip=True)
            prize_val = self._parse_amount(prize_text)

            # Total Winners (last col or dedicated col)
            if total_winners_col is not None and total_winners_col < len(cells):
                winners_text = cells[total_winners_col].get_text(strip=True)
            else:
                winners_text = cells[-1].get_text(strip=True)
            # Strip "Rollover!" prefix from winners text
            winners_text = re.sub(r"(?i)rollover!?\s*", "", winners_text).strip()
            winners = self._parse_int(winners_text)

            result[f"t{tier}_prize"] = prize_val
            result[f"t{tier}_winners"] = winners

            # Jackpot is tier 1 prize
            if tier == 1:
                result["jackpot"] = prize_val

        return result if result else None

    @staticmethod
    def _parse_amount(text: str) -> float:
        """Parse '€1,234,567.89' or '1.234.567,89 €' into float."""
        import re
        text = text.replace("€", "").replace("£", "").replace("$", "").strip()
        if not text or text.lower() in ("rollover", "rolldown", "-", "n/a"):
            return 0.0
        # Detect European format: 1.234.567,89
        if "," in text and "." in text:
            if text.rindex(",") > text.rindex("."):
                # European: dots are thousands, comma is decimal
                text = text.replace(".", "").replace(",", ".")
            else:
                # US/UK: commas are thousands, dot is decimal
                text = text.replace(",", "")
        elif "," in text and "." not in text:
            # Could be decimal comma (e.g. "17,08") or thousands (e.g. "1,234")
            parts = text.split(",")
            if len(parts) == 2 and len(parts[1]) == 2:
                text = text.replace(",", ".")
            else:
                text = text.replace(",", "")
        try:
            return float(re.sub(r"[^\d.]", "", text))
        except ValueError:
            return 0.0

    @staticmethod
    def _parse_int(text: str) -> int:
        """Parse '1,234' or '1.234' into int."""
        import re
        text = re.sub(r"[^\d]", "", text)
        try:
            return int(text)
        except ValueError:
            return 0

    def scrape_premios_bulk(self, datas: list[str], db: "DatabaseManager"):
        """Generator: scrape prizes for multiple dates, yielding progress events."""
        total = len(datas)
        yield {"tipo": "inicio", "total": total}

        sucesso = 0
        falha = 0
        for i, data in enumerate(datas):
            yield {"tipo": "progresso", "step": i + 1, "total": total, "data": data}
            try:
                prize_data = self.scrape_premios(data)
                if prize_data:
                    db.inserir_premios(data, prize_data)
                    sucesso += 1
                    yield {"tipo": "ok", "data": data, "step": i + 1}
                else:
                    falha += 1
                    yield {"tipo": "sem_dados", "data": data, "step": i + 1}
            except Exception as e:
                falha += 1
                yield {"tipo": "erro", "data": data, "msg": str(e)[:120], "step": i + 1}
            time.sleep(SCRAPE_DELAY)  # Rate limit

        yield {"tipo": "concluido", "sucesso": sucesso, "falha": falha, "total": total}


# ═════════════════════════════════════════════════════════════════════════════
# WEB SCRAPER
# ═════════════════════════════════════════════════════════════════════════════
class EuromilhoesScraper:
    HEADERS = {
        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
    }

    def fetch_ultimo_sorteio(self) -> dict | None:
        """Try multiple sources to get the latest EuroMillions draw."""
        strategies = [
            self._scrape_lotaria_net,
            self._scrape_national_lottery_api,
            self._scrape_euro_jackpot_results,
        ]
        for strategy in strategies:
            try:
                result = strategy()
                if result:
                    return result
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                continue
            except Exception:
                continue
        return None

    def _find_date_in_soup(self, soup) -> str | None:
        """Try to find a draw date anywhere in the page."""
        # Try <time> elements first
        for t in soup.find_all("time"):
            raw = t.get("datetime") or t.get_text(strip=True)
            d = self._try_parse_date(raw)
            if d:
                return d
        # Try elements with date-related class names
        for elem in soup.find_all(["span", "div", "p", "h2", "h3", "h4"],
                                  class_=lambda c: c and any(
                                      w in str(c).lower() for w in ["date", "data", "draw-date"]
                                  )):
            d = self._try_parse_date(elem.get_text(strip=True))
            if d:
                return d
        # Scan all text for a date pattern
        import re
        for text in soup.stripped_strings:
            m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', text)
            if m:
                return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
            m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', text)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return None

    @staticmethod
    def _try_parse_date(text: str) -> str | None:
        """Try to parse a date string, return ISO format or None."""
        if not text or len(text) < 6:
            return None
        import re
        m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', text)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', text)
        if m:
            return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
        return None

    def _fallback_draw_date(self) -> str:
        """When no date can be scraped, snap today to the most recent Tue/Fri."""
        return corrigir_data_sorteio(datetime.date.today().isoformat())

    def _scrape_lotaria_net(self) -> dict | None:
        url = "https://www.lotaria.net/euromilhoes/resultados"
        r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        date_str = self._find_date_in_soup(soup) or self._fallback_draw_date()

        # Try common result container patterns
        selectors = [
            ("div", {"class": lambda c: c and "resultado" in c.lower()}),
            ("div", {"class": lambda c: c and "result" in c.lower()}),
            ("ul", {"class": lambda c: c and "ball" in c.lower()}),
        ]
        for tag, attrs in selectors:
            container = soup.find(tag, attrs)
            if container:
                nums = self._extract_numbers_from_container(container, date_str)
                if nums:
                    return nums

        # Fallback: find all number balls
        balls = soup.find_all(["span", "li", "div"], class_=lambda c: c and (
            "ball" in c.lower() or "numero" in c.lower() or "number" in c.lower()
        ))
        all_nums = []
        for b in balls:
            txt = b.get_text(strip=True)
            if txt.isdigit():
                all_nums.append(int(txt))

        return self._parse_raw_numbers(all_nums, date_str)

    def _scrape_national_lottery_api(self) -> dict | None:
        # Try the unofficial API used by some Portuguese lottery sites
        urls = [
            "https://www.jogossantacasa.pt/web/SCCartazResult/euroMilhoes",
            "https://api.lotaria.net/api/euromilhoes/last",
        ]
        for url in urls:
            try:
                r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
                if r.status_code == 200:
                    # Try JSON first
                    try:
                        data = r.json()
                        return self._parse_json_result(data)
                    except (ValueError, KeyError):
                        pass
                    # Try HTML
                    soup = BeautifulSoup(r.text, "html.parser")
                    date_str = self._find_date_in_soup(soup) or self._fallback_draw_date()
                    balls = soup.find_all(class_=lambda c: c and "ball" in str(c).lower())
                    nums = []
                    for b in balls:
                        t = b.get_text(strip=True)
                        if t.isdigit():
                            nums.append(int(t))
                    result = self._parse_raw_numbers(nums, date_str)
                    if result:
                        return result
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                continue
            except Exception:
                continue
        return None

    def _scrape_euro_jackpot_results(self) -> dict | None:
        """Fallback: try eurojackpot-style API with EuroMillions data."""
        url = "https://www.euro-millions.com/results"
        r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        date_str = self._find_date_in_soup(soup) or self._fallback_draw_date()

        balls = soup.find_all(["li", "span", "div"], class_=lambda c: c and (
            "ball" in str(c).lower() or "num" in str(c).lower()
        ))
        nums = []
        for b in balls:
            t = b.get_text(strip=True)
            if t.isdigit():
                nums.append(int(t))

        return self._parse_raw_numbers(nums, date_str)

    def _extract_numbers_from_container(self, container, date_str: str = None) -> dict | None:
        texts = []
        for elem in container.find_all(text=True):
            txt = elem.strip()
            if txt.isdigit():
                texts.append(int(txt))
        return self._parse_raw_numbers(texts, date_str)

    def _parse_raw_numbers(self, nums: list, date_str: str = None) -> dict | None:
        """Given a flat list of numbers, try to identify 5 main + 2 stars."""
        if not date_str:
            date_str = self._fallback_draw_date()

        main_candidates = [n for n in nums if 1 <= n <= 50]
        star_candidates = [n for n in nums if 1 <= n <= 12]

        # Need exactly 5 unique main numbers and 2 unique stars
        main_nums = list(dict.fromkeys(main_candidates))[:5]
        star_nums = list(dict.fromkeys(star_candidates))[:2]

        if len(main_nums) == 5 and len(star_nums) == 2:
            return {"data": date_str, "numeros": sorted(main_nums), "estrelas": sorted(star_nums)}
        return None

    def _parse_json_result(self, data: dict) -> dict | None:
        """Try common JSON field names."""
        try:
            # Various APIs use different field names
            for nums_key in ["numbers", "numeros", "balls", "mainNumbers"]:
                for stars_key in ["stars", "estrelas", "luckyStars", "bonusBalls"]:
                    if nums_key in data and stars_key in data:
                        nums = sorted([int(n) for n in data[nums_key]])[:5]
                        stars = sorted([int(s) for s in data[stars_key]])[:2]
                        date_str = data.get("date", data.get("data", self._fallback_draw_date()))
                        if len(nums) == 5 and len(stars) == 2:
                            return {"data": str(date_str), "numeros": nums, "estrelas": stars}
        except (ValueError, KeyError, TypeError):
            pass
        return None

    def fetch_historico(self, max_draws: int = 200) -> list[dict]:
        """Try to fetch a batch of historical results."""
        results = []
        urls_to_try = [
            f"https://www.lotaria.net/euromilhoes/arquivo?page=1",
            f"https://www.euro-millions.com/results/2024",
            f"https://www.euro-millions.com/results/2023",
        ]
        for url in urls_to_try:
            try:
                r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
                if r.status_code != 200:
                    continue
                soup = BeautifulSoup(r.text, "html.parser")
                # Look for result rows
                rows = soup.find_all(["tr", "div", "article"], class_=lambda c: c and (
                    "result" in str(c).lower() or "draw" in str(c).lower() or "sorteio" in str(c).lower()
                ))
                for row in rows[:max_draws]:
                    nums_raw = []
                    for elem in row.find_all(text=True):
                        t = elem.strip()
                        if t.isdigit():
                            nums_raw.append(int(t))
                    parsed = self._parse_raw_numbers(nums_raw)
                    if parsed:
                        results.append(parsed)
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                continue
            except Exception:
                continue
        return results


# ═════════════════════════════════════════════════════════════════════════════
# HISTORICAL BULK SCRAPER
# ═════════════════════════════════════════════════════════════════════════════
import re as _re

class HistoricoScraper:
    """
    Scrapes the COMPLETE EuroMillions history year by year (2004 → present).
    Uses a generator so progress can be streamed to the caller.

    Yield dict types:
      {"tipo": "inicio",       "total_anos": N}
      {"tipo": "progresso",    "ano": Y, "step": i, "total": N}
      {"tipo": "ano_ok",       "ano": Y, "encontrados": k, "total_acumulado": T}
      {"tipo": "ano_erro",     "ano": Y, "msg": "..."}
      {"tipo": "concluido",    "total": T, "inseridos_db": I, "ficheiro": "..."}
    """

    START_YEAR = 2004   # EuroMillions first draw: 7 Feb 2004
    DELAY      = 1.2    # seconds between HTTP requests (be polite)

    HEADERS = {
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-GB,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
    }

    MONTHS = {
        "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
        "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
        "january":1,"february":2,"march":3,"april":4,"june":6,"july":7,
        "august":8,"september":9,"october":10,"november":11,"december":12,
        "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
        "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12,
    }

    def scrape_completo(self, db: "DatabaseManager", ficheiro: Path):
        """Generator – call with `for evento in scraper.scrape_completo(db, path):`"""
        ano_atual  = datetime.date.today().year
        anos       = list(range(self.START_YEAR, ano_atual + 1))
        acumulado  = []

        yield {"tipo": "inicio", "total_anos": len(anos)}

        for i, ano in enumerate(anos):
            yield {"tipo": "progresso", "ano": ano, "step": i + 1, "total": len(anos)}

            try:
                resultados = self._scrape_ano(ano)
                acumulado.extend(resultados)
                # Persist partial results after every year
                self._salvar_ficheiro(ficheiro, acumulado)
                yield {"tipo": "ano_ok", "ano": ano,
                       "encontrados": len(resultados), "total_acumulado": len(acumulado)}
            except Exception as exc:
                yield {"tipo": "ano_erro", "ano": ano, "msg": str(exc)[:120]}

            time.sleep(self.DELAY)

        # Final DB import
        inseridos = 0
        for s in acumulado:
            if db.inserir_sorteio(s["data"], s["numeros"], s["estrelas"], "historico"):
                inseridos += 1

        yield {"tipo": "concluido", "total": len(acumulado),
               "inseridos_db": inseridos, "ficheiro": str(ficheiro)}

    def scrape_desde(self, desde_data: str, db: "DatabaseManager") -> dict:
        """Scrape only the draws missing between last DB date and today."""
        existentes = db.todos_sorteios()
        datas_existentes = {s["data"] for s in existentes}

        # Enumerate expected draw dates (Tue/Fri) between last DB entry and today
        desde = datetime.date.fromisoformat(desde_data)
        hoje = datetime.date.today()
        datas_em_falta = []
        d = desde + datetime.timedelta(days=1)
        while d <= hoje:
            if d.weekday() in (1, 4) and d.isoformat() not in datas_existentes:
                datas_em_falta.append(d)
            d += datetime.timedelta(days=1)

        if not datas_em_falta:
            return {"encontrados": 0, "inseridos": 0, "sorteios": []}

        todos = []
        if len(datas_em_falta) <= 5:
            # Few missing: fetch each individual draw page
            for data in datas_em_falta:
                url = f"https://www.euro-millions.com/results/{data.day:02d}-{data.month:02d}-{data.year}"
                try:
                    r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
                    if r.status_code == 200:
                        resultados = self._parse_results_table(r.text)
                        if resultados:
                            s = resultados[0]
                            s["data"] = data.isoformat()
                            todos.append(s)
                except Exception:
                    continue
        else:
            # Many missing: fetch full year history pages (1-2 requests)
            anos = sorted({d.year for d in datas_em_falta})
            for ano in anos:
                url = f"https://www.euro-millions.com/results-history-{ano}"
                try:
                    r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
                    if r.status_code == 200:
                        todos.extend(self._parse_results_table(r.text))
                except Exception:
                    continue

        # Filter to only new draws
        vistos = set()
        unicos = []
        for s in todos:
            if s["data"] not in datas_existentes and s["data"] not in vistos:
                vistos.add(s["data"])
                unicos.append(s)
        unicos.sort(key=lambda s: s["data"])

        inseridos = 0
        for s in unicos:
            if db.inserir_sorteio(s["data"], s["numeros"], s["estrelas"], "web-auto"):
                inseridos += 1

        return {"encontrados": len(unicos), "inseridos": inseridos, "sorteios": unicos}

    # ── Per-year scraping ─────────────────────────────────────────────────────
    def _scrape_ano(self, ano: int) -> list[dict]:
        for fn in [self._euro_millions_com, self._euro_millions_com_pt,
                   self._lotaria_net_arquivo]:
            try:
                resultados = fn(ano)
                if resultados:
                    return resultados
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
                continue
            except Exception:
                continue
        return []

    def _euro_millions_com_recent(self) -> list[dict]:
        """Scrape results-history pages for current and previous year."""
        ano_atual = datetime.date.today().year
        todos = []
        for ano in [ano_atual, ano_atual - 1]:
            url = f"https://www.euro-millions.com/results-history-{ano}"
            try:
                r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
                if r.status_code == 200:
                    todos.extend(self._parse_results_table(r.text))
            except Exception:
                continue
        return self._deduplicate(todos)

    def _parse_results_table(self, html: str) -> list[dict]:
        """Parse euro-millions.com/results-history-YYYY table format."""
        soup = BeautifulSoup(html, "html.parser")
        results = []
        for ul in soup.find_all("ul", class_="balls"):
            balls = ul.find_all("li", class_="resultBall")
            main_nums = [int(b.get_text(strip=True)) for b in balls
                         if "ball" in " ".join(b.get("class", [])) and "lucky-star" not in " ".join(b.get("class", []))]
            star_nums = [int(b.get_text(strip=True)) for b in balls
                         if "lucky-star" in " ".join(b.get("class", []))]
            if len(main_nums) < 5 or len(star_nums) < 2:
                continue
            date_str = None
            tr = ul.find_parent("tr")
            if tr:
                link = tr.find("a", href=_re.compile(r"/results/(\d{2})-(\d{2})-(\d{4})"))
                if link:
                    m = _re.search(r"(\d{2})-(\d{2})-(\d{4})", link["href"])
                    if m:
                        date_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            if date_str:
                results.append({
                    "data": date_str,
                    "numeros": sorted(main_nums[:5]),
                    "estrelas": sorted(star_nums[:2]),
                })
        return results

    def _euro_millions_com(self, ano: int) -> list[dict]:
        url = f"https://www.euro-millions.com/results-history-{ano}"
        try:
            r = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
            if r.status_code != 200:
                return []
            return self._parse_results_table(r.text)
        except Exception:
            return []

    def _euro_millions_com_pt(self, ano: int) -> list[dict]:
        url = f"https://www.euro-millions.com/pt/resultados/{ano}"
        r   = _retry_request(url, headers=self.HEADERS, timeout=REQUEST_TIMEOUT)
        if r.status_code != 200:
            return []
        return self._parse_html(r.text, ano)

    def _lotaria_net_arquivo(self, ano: int) -> list[dict]:
        url = f"https://www.lotaria.net/euromilhoes/arquivo/{ano}"
        r   = _retry_request(url, headers={**self.HEADERS, "Accept-Language": "pt-PT,pt;q=0.9"},
                             timeout=REQUEST_TIMEOUT)
        if r.status_code != 200:
            return []
        return self._parse_html(r.text, ano)

    # ── HTML parsing ──────────────────────────────────────────────────────────
    def _parse_html(self, html: str, ano: int) -> list[dict]:
        soup    = BeautifulSoup(html, "html.parser")
        results = []

        # Remove noise
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()

        # Strategy A – find result containers with balls
        containers = soup.find_all(
            ["article", "section", "div", "li"],
            class_=lambda c: c and any(
                w in str(c).lower() for w in
                ["result", "draw", "lottery", "sorteio", "resultado"]
            ),
        )
        for cont in containers:
            d = self._parse_container(cont, ano)
            if d:
                results.append(d)

        # Strategy B – table rows
        if not results:
            for table in soup.find_all("table"):
                for row in table.find_all("tr"):
                    d = self._parse_table_row(row, ano)
                    if d:
                        results.append(d)

        # Strategy C – aggressive text regex fallback
        if not results:
            results = self._regex_fallback(soup.get_text(" "), ano)

        return self._deduplicate(results)

    def _parse_container(self, cont, ano: int) -> dict | None:
        # ── Date ──
        date_str = None
        for elem in cont.find_all(["time", "span", "a", "h2", "h3", "h4", "p", "div"]):
            raw = elem.get("datetime") or elem.get_text(" ", strip=True)
            date_str = self._parse_date(raw, ano)
            if date_str:
                break
        if not date_str:
            return None

        # ── Main numbers ──
        main_nums = []
        star_nums = []

        # Try to find separate star container first
        star_cont = cont.find(
            class_=lambda c: c and any(w in str(c).lower() for w in ["star", "lucky", "bonus"])
        )

        ball_elems = cont.find_all(
            ["li", "span", "div", "td"],
            class_=lambda c: c and any(
                w in str(c).lower() for w in ["ball", "num", "number", "main", "boule"]
            ),
        )
        for b in ball_elems:
            t = b.get_text(strip=True)
            if t.isdigit():
                n = int(t)
                if star_cont and b in star_cont.descendants:
                    if 1 <= n <= 12:
                        star_nums.append(n)
                elif 1 <= n <= 50:
                    main_nums.append(n)

        # If we couldn't distinguish, use range heuristic:
        # after 5 main numbers, remaining small numbers are stars
        if len(main_nums) >= 7 and not star_nums:
            all_nums = sorted(set(main_nums))
            main_nums = all_nums[:5]
            star_nums = [n for n in all_nums[5:] if 1 <= n <= 12][:2]

        if len(main_nums) >= 5 and len(star_nums) >= 2:
            return {"data": date_str,
                    "numeros": sorted(set(main_nums))[:5],
                    "estrelas": sorted(set(star_nums))[:2]}
        return None

    def _parse_table_row(self, row, ano: int) -> dict | None:
        cells = row.find_all(["td", "th"])
        if len(cells) < 3:
            return None
        nums_in_row = []
        date_str    = None
        for cell in cells:
            raw = cell.get_text(" ", strip=True)
            if not date_str:
                date_str = self._parse_date(raw, ano)
            if raw.isdigit():
                nums_in_row.append(int(raw))
        if len(nums_in_row) >= 7 and date_str:
            sorted_nums = sorted(set(nums_in_row))
            main   = [n for n in sorted_nums if 1 <= n <= 50][:5]
            stars  = [n for n in sorted_nums if 1 <= n <= 12][:2]
            if len(main) == 5 and len(stars) == 2:
                return {"data": date_str, "numeros": main, "estrelas": stars}
        return None

    def _regex_fallback(self, text: str, ano: int) -> list[dict]:
        """
        Last resort: scan plain text for lines that contain a date
        followed by 7 space/punct-separated numbers.
        """
        results = []
        # Find every sequence of 7+ integers on the same "block"
        pattern = _re.compile(
            r'(\d{1,2}[./ -]\d{1,2}[./ -]\d{4}|\d{4}[./ -]\d{2}[./ -]\d{2})'
            r'[^\d]*'
            r'(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})'
            r'\D+(\d{1,2})\D+(\d{1,2})'
        )
        for m in pattern.finditer(text):
            date_str = self._parse_date(m.group(1), ano)
            if not date_str:
                continue
            nums = [int(m.group(i)) for i in range(2, 9)]
            main  = sorted([n for n in nums if 1 <= n <= 50])
            stars = sorted([n for n in nums if 1 <= n <= 12])
            if len(main) >= 5 and len(stars) >= 2:
                results.append({"data": date_str,
                                 "numeros": main[:5], "estrelas": stars[:2]})
        return results

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _parse_date(self, text: str, fallback_year: int = None) -> str | None:
        if not text or len(text) < 4:
            return None
        text = text.strip()

        # ISO: YYYY-MM-DD
        m = _re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', text)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

        # DD/MM/YYYY or DD.MM.YYYY or DD-MM-YYYY
        m = _re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', text)
        if m:
            return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"

        # "Friday 12th January 2024" or "12 January 2024"
        m = _re.search(r'(\d{1,2})\w*\s+([A-Za-z]+)\s+(\d{4})', text)
        if m:
            mon = self.MONTHS.get(m.group(2).lower()[:9])
            if mon:
                return f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}"

        # "January 12, 2024"
        m = _re.search(r'([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})', text)
        if m:
            mon = self.MONTHS.get(m.group(1).lower()[:9])
            if mon:
                return f"{m.group(3)}-{mon:02d}-{int(m.group(2)):02d}"

        return None

    def _deduplicate(self, results: list[dict]) -> list[dict]:
        seen = set()
        out  = []
        for r in results:
            key = r["data"]
            if key not in seen:
                seen.add(key)
                out.append(r)
        return sorted(out, key=lambda x: x["data"])

    @staticmethod
    def _salvar_ficheiro(ficheiro: Path, sorteios: list[dict]):
        data = {
            "gerado_em": datetime.date.today().isoformat(),
            "total":     len(sorteios),
            "sorteios":  sorteios,
        }
        ficheiro.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


if _IS_VERCEL:
    HISTORICO_PATH    = Path("/tmp/historico_completo.json")
    EXCEL_SOURCE_PATH = Path("/tmp/Euromilhões _ Todos os sorteios.xlsx")
else:
    HISTORICO_PATH    = Path(__file__).parent / "historico_completo.json"
    EXCEL_SOURCE_PATH = Path(__file__).parent / "Euromilhões _ Todos os sorteios.xlsx"


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL IMPORTER
# ═════════════════════════════════════════════════════════════════════════════
import re as _re2

class ExcelImporter:
    """
    Imports the complete draw history from the local Excel file.

    File:  'Euromilhões _ Todos os sorteios.xlsx'
    Sheet: 'EUROMILHÕES-Zeros'
    Layout (1-based columns):
      Col B (2)  : draw sequence number
      Col C (3)  : date  (datetime object or string like '\xa013\xa0fev\xa02004')
      Cols D–H (4–8) : 5 drawn numbers
      Cols I–J (9–10): 2 lucky stars
    Data starts at row 14 (row 13 is the upcoming-draw placeholder).
    """

    SHEET_NAME    = "EUROMILHÕES-Zeros"
    DATA_START_ROW = 14   # row 13 = upcoming draw placeholder → skip

    PT_MONTHS = {
        "jan": 1, "fev": 2, "mar": 3, "abr": 4,
        "mai": 5, "jun": 6, "jul": 7, "ago": 8,
        "set": 9, "out": 10, "nov": 11, "dez": 12,
    }

    def __init__(self, excel_path: Path = None):
        self.excel_path = excel_path or EXCEL_SOURCE_PATH

    # ── Public API ────────────────────────────────────────────────────────────
    def importar(self, db: "DatabaseManager") -> dict:
        """Read the Excel file and insert all draws.  Returns summary dict."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Ficheiro não encontrado: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
        try:
            ws = wb[self.SHEET_NAME]
        except KeyError:
            wb.close()
            raise ValueError(f"Folha '{self.SHEET_NAME}' não encontrada.")

        inseridos = ja_existiam = erros = total_lido = 0

        for row in ws.iter_rows(min_row=self.DATA_START_ROW, values_only=True):
            draw_num              = row[1]   # Col B – draw sequence number
            date_val              = row[2]   # Col C – date
            n1, n2, n3, n4, n5   = row[3], row[4], row[5], row[6], row[7]
            e1, e2               = row[8], row[9]

            if draw_num is None:             # past end of data
                break

            data_str = self._parse_date(date_val)
            if data_str is None:
                erros += 1
                continue

            try:
                nums  = sorted([int(n1), int(n2), int(n3), int(n4), int(n5)])
                stars = sorted([int(e1), int(e2)])
                assert len(set(nums))  == 5 and all(1 <= n <= 50 for n in nums)
                assert len(set(stars)) == 2 and all(1 <= s <= 12 for s in stars)
            except (TypeError, ValueError, AssertionError):
                erros += 1
                continue

            total_lido += 1
            ok = db.inserir_sorteio(data_str, nums, stars, fonte="excel")
            if ok:
                inseridos += 1
            else:
                ja_existiam += 1

        wb.close()
        return {
            "inseridos":    inseridos,
            "ja_existiam":  ja_existiam,
            "erros":        erros,
            "total_lido":   total_lido,
        }

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _parse_date(self, val) -> str | None:
        if val is None:
            return None
        if isinstance(val, (datetime.datetime, datetime.date)):
            if isinstance(val, datetime.datetime):
                return val.date().isoformat()
            return val.isoformat()
        # String e.g. '\xa013\xa0fev\xa02004'
        text = str(val).replace("\xa0", " ").strip()
        parts = text.split()
        if len(parts) >= 3:
            try:
                day   = int(parts[0])
                month = self.PT_MONTHS.get(parts[1].lower()[:3])
                year  = int(parts[2])
                if month and 1 <= day <= 31 and 2004 <= year <= 2030:
                    return f"{year}-{month:02d}-{day:02d}"
            except (ValueError, IndexError):
                pass
        m = _re2.search(r"(\d{4})-(\d{2})-(\d{2})", text)
        if m:
            return m.group(0)
        return None


# ═════════════════════════════════════════════════════════════════════════════
# STATISTICS ANALYSER
# ═════════════════════════════════════════════════════════════════════════════
class StatisticsAnalyzer:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def classificar_cores(self, ultimos_9: list[dict]) -> dict:
        """
        Returns {"vermelhos": set, "verdes": set, "azuis": set, "castanhos": set}
        based on how many times each number 1–50 appeared in the last 9 draws.
        """
        contagem = Counter()
        for s in ultimos_9:
            for n in s["numeros"]:
                contagem[n] += 1

        vermelhos, verdes, azuis, castanhos = set(), set(), set(), set()
        for n in range(1, 51):
            c = contagem[n]
            if c == 0:
                vermelhos.add(n)
            elif c == 1:
                verdes.add(n)
            elif c == 2:
                azuis.add(n)
            else:  # 3+
                castanhos.add(n)

        return {"vermelhos": vermelhos, "verdes": verdes,
                "azuis": azuis, "castanhos": castanhos}

    def frequencia_numeros(self) -> dict:
        todos = self.db.todos_sorteios()
        freq = Counter()
        for s in todos:
            for n in s["numeros"]:
                freq[n] += 1
        return dict(freq)

    def frequencia_estrelas(self) -> dict:
        todos = self.db.todos_sorteios()
        freq = Counter()
        for s in todos:
            for e in s["estrelas"]:
                freq[e] += 1
        return dict(freq)

    def estatisticas_somas(self) -> dict:
        todos = self.db.todos_sorteios()
        somas = [s["soma"] for s in todos if s["soma"]]
        if not somas:
            return {}
        return {
            "min": min(somas), "max": max(somas),
            "media": round(sum(somas) / len(somas), 1),
            "mais_comum": Counter(somas).most_common(5),
        }

    def analise_padroes(self) -> dict:
        todos = self.db.todos_sorteios()
        padrao_count = Counter()
        for s in todos:
            nums = s["numeros"]
            bi_c = sum(1 for n in nums if n in BI)
            bp_c = sum(1 for n in nums if n in BP)
            ai_c = sum(1 for n in nums if n in AI)
            ap_c = sum(1 for n in nums if n in AP)
            padrao = tuple([bi_c, bp_c, ai_c, ap_c])
            padrao_count[padrao] += 1
        return dict(padrao_count.most_common(16))

    def numeros_atrasados(self, n: int = 10) -> list:
        """Numbers that have not appeared for the longest time."""
        todos = self.db.todos_sorteios()
        ultima_vez = {}
        for i, s in enumerate(reversed(todos)):  # oldest first
            for num in s["numeros"]:
                ultima_vez[num] = i
        ausentes = []
        for num in range(1, 51):
            draws_ago = len(todos) - ultima_vez.get(num, -1) - 1
            ausentes.append((num, draws_ago))
        ausentes.sort(key=lambda x: x[1], reverse=True)
        return ausentes[:n]

    def sequencias_quentes(self, janela: int = 10) -> list:
        """Numbers most frequent in the last `janela` draws."""
        ultimos = self.db.ultimos_n_sorteios(janela)
        freq = Counter()
        for s in ultimos:
            for n in s["numeros"]:
                freq[n] += 1
        return freq.most_common(10)

    def quentes_frios_completo(self, janela: int = 15) -> list[dict]:
        """For every number 1-50: recent freq, total freq, draws since last, avg gap, due flag."""
        todos = self.db.todos_sorteios()  # newest first
        total = len(todos)
        ultimos = todos[:janela]

        freq_recente = Counter()
        for s in ultimos:
            for n in s["numeros"]:
                freq_recente[n] += 1

        freq_total = Counter()
        for s in todos:
            for n in s["numeros"]:
                freq_total[n] += 1

        result = []
        for num in range(1, 51):
            aparicoes = [i for i, s in enumerate(todos) if num in s["numeros"]]
            draws_since = aparicoes[0] if aparicoes else total

            if len(aparicoes) >= 2:
                gaps = [aparicoes[j + 1] - aparicoes[j] for j in range(len(aparicoes) - 1)]
                avg_gap = round(sum(gaps) / len(gaps), 1)
            else:
                avg_gap = float(total)

            result.append({
                "num": num,
                "freq_recente": freq_recente.get(num, 0),
                "freq_total": freq_total.get(num, 0),
                "draws_since": draws_since,
                "avg_gap": avg_gap,
                "due": draws_since > avg_gap * 1.5 if avg_gap > 0 else False,
            })
        return result

    def analise_gaps(self) -> list[dict]:
        """For each number 1-50: current gap, avg gap, max gap, ratio, sorted by ratio desc."""
        todos = self.db.todos_sorteios()  # newest first
        total = len(todos)

        result = []
        for num in range(1, 51):
            aparicoes = [i for i, s in enumerate(todos) if num in s["numeros"]]
            current_gap = aparicoes[0] if aparicoes else total

            if len(aparicoes) >= 2:
                gaps = [aparicoes[j + 1] - aparicoes[j] for j in range(len(aparicoes) - 1)]
                avg_gap = round(sum(gaps) / len(gaps), 1)
                max_gap = max(gaps)
            else:
                avg_gap = float(total)
                max_gap = total

            ratio = round(current_gap / avg_gap, 2) if avg_gap > 0 else 0
            result.append({
                "num": num,
                "current_gap": current_gap,
                "avg_gap": avg_gap,
                "max_gap": max_gap,
                "ratio": ratio,
                "total_aparicoes": len(aparicoes),
            })
        result.sort(key=lambda x: x["ratio"], reverse=True)
        return result

    def tendencia_somas(self, janela: int = 30) -> list[dict]:
        """Sum per draw + running average for the last `janela` draws (chronological order)."""
        todos = self.db.todos_sorteios()[:janela]  # newest first
        todos.reverse()  # oldest first

        result = []
        running = 0
        for i, s in enumerate(todos):
            soma = s.get("soma") or sum(s["numeros"])
            running += soma
            result.append({
                "data": s["data"],
                "soma": soma,
                "media_movel": round(running / (i + 1), 1),
            })
        return result

    def historico_padroes(self) -> dict:
        """Count how many historical draws match each of the 56 BI-BP-AI-AP patterns."""
        todos = self.db.todos_sorteios()
        total = len(todos)
        contagem: dict[int, int] = {}
        for s in todos:
            p = classificar_padrao_bibpaiap(s["numeros"])
            if p:
                contagem[p["num"]] = contagem.get(p["num"], 0) + 1

        # Group summaries
        _grupos = [
            {"chave": "best",        "nome": "Melhores (4)",       "nums": range(1,  5),  "pct_esp": 26.42},
            {"chave": "middle_high", "nome": "Médio/Alto (12)",     "nums": range(5,  17), "pct_esp": 36.53},
            {"chave": "middle_low",  "nome": "Médio/Baixo (12)",    "nums": range(17, 29), "pct_esp": 22.30},
            {"chave": "worst",       "nome": "Baixo/Mto Baixo (28)","nums": range(29, 57), "pct_esp": 14.75},
        ]
        resumo_grupos = []
        for g in _grupos:
            real = sum(contagem.get(i, 0) for i in g["nums"])
            esperado = round(total * g["pct_esp"] / 100)
            resumo_grupos.append({
                "nome":       g["nome"],
                "chave":      g["chave"],
                "esperado":   esperado,
                "pct_esp":    g["pct_esp"],
                "real":       real,
                "pct_real":   round(real / total * 100, 2) if total > 0 else 0,
            })

        # Per-pattern detail
        detalhe = []
        for p in TODOS_PADROES_BIBPAIAP:
            real = contagem.get(p["num"], 0)
            detalhe.append({
                "num":         p["num"],
                "padrao":      p["padrao"],
                "combinacoes": p["combinacoes"],
                "pct":         p["pct"],
                "grupo":       p["grupo"],
                "real":        real,
                "pct_real":    round(real / total * 100, 2) if total > 0 else 0,
            })

        return {"total": total, "resumo_grupos": resumo_grupos, "detalhe": detalhe}


# ═════════════════════════════════════════════════════════════════════════════
# FILTER ENGINE
# ═════════════════════════════════════════════════════════════════════════════
class FilterEngine:
    """
    All filters ordered by computational cost (cheapest first).
    All filters A-H can be independently toggled via config.

    Config keys (all optional, all default to True/active):
      soma_range : "padrao" (80–190) | "apertado" (95–160)
      filtro_A   : bool – Soma range
      filtro_B   : bool – Consecutive pairs/triples
      filtro_C   : bool – Same final digit
      filtro_D   : bool – Decades spread
      filtro_E   : bool – Repeat from last draw
      filtro_F   : bool – Colour system (last 9 draws)
      filtro_G   : bool – Regra do 31 (also accepts legacy key "regra31")
      filtro_H   : bool – Anti-arithmetic-progression (also accepts legacy "progressao")

    To add a new filter: add its id to ALL_FILTERS, add info to FILTER_INFO,
    add stat key to stats dict, add check block in verificar(), add to STAT_MAP.
    """

    ALL_FILTERS = list("ABCDEFGH")

    FILTER_INFO = {
        "A": {
            "nome": "Soma",
            "descricao": "Limita a soma dos 5 números ao intervalo configurado. "
                         "Cerca de 93% dos sorteios históricos têm somas entre 80–190.",
        },
        "B": {
            "nome": "Consecutivos",
            "descricao": "Permite no máximo 1 par de números consecutivos e proíbe triplos. "
                         "Pares ocorrem em ~42% dos sorteios; triplos em menos de 1%.",
        },
        "C": {
            "nome": "Dígito Final",
            "descricao": "Máx 2 números com o mesmo dígito final (ex: 3, 13, 23). "
                         "Três ou mais iguais representam menos de 4% dos sorteios.",
        },
        "D": {
            "nome": "Dezenas",
            "descricao": "Mín 3 dezenas diferentes (1-10, 11-20, 21-30, 31-40, 41-50). "
                         "Evita concentração de números numa só zona do boletim.",
        },
        "E": {
            "nome": "Anti-Repetição",
            "descricao": "Máx 2 números em comum com o sorteio anterior. "
                         "Baseado em tendências de curto prazo observadas no histórico.",
        },
        "F": {
            "nome": "Cores (9 sorteios)",
            "descricao": "Sistema de cores baseado nos últimos 9 sorteios: "
                         "Vermelhos (0×) → 1-3 por chave; Verdes (1×) → 1-3; "
                         "Azuis (2×) → 0-2; Castanhos (3×+) → excluídos.",
        },
        "G": {
            "nome": "Regra do 31",
            "descricao": "Mín 1 número acima de 31. A maioria dos jogadores usa aniversários "
                         "(1-31), forçar números altos reduz a partilha do jackpot.",
        },
        "H": {
            "nome": "Anti-Progressão",
            "descricao": "Rejeita progressões aritméticas perfeitas (ex: 5,10,15,20,25). "
                         "Estas sequências são muito populares entre jogadores ingénuos.",
        },
    }

    STAT_MAP = {
        "A": "reprovadas_soma",
        "B": "reprovadas_consecutivos",
        "C": "reprovadas_finais",
        "D": "reprovadas_decadas",
        "E": "reprovadas_repeticao",
        "F": "reprovadas_cores",
        "G": "reprovadas_regra31",
        "H": "reprovadas_progressao",
    }

    SOMA_RANGES = {
        "padrao":   (80, 190),
        "apertado": (95, 160),
    }

    def __init__(self, cores: dict, ultimo_sorteio: list | None = None,
                 config: dict | None = None):
        self.cores = cores
        self.ultimo_sorteio = ultimo_sorteio or []
        cfg = config or {}

        rng = self.SOMA_RANGES.get(cfg.get("soma_range", "padrao"), (80, 190))
        self.soma_min = cfg.get("soma_min", rng[0])
        self.soma_max = cfg.get("soma_max", rng[1])

        # Build active_filters set – all enabled by default.
        # Backward compat: G also accepts "regra31", H accepts "progressao".
        self.active_filters: set[str] = set()
        for f in self.ALL_FILTERS:
            if f == "G":
                default = cfg.get("regra31", True)
            elif f == "H":
                default = cfg.get("progressao", True)
            else:
                default = True
            if cfg.get(f"filtro_{f}", default):
                self.active_filters.add(f)

        self.stats: dict[str, int] = {
            "testadas": 0, "aprovadas": 0,
            "reprovadas_soma": 0, "reprovadas_consecutivos": 0,
            "reprovadas_finais": 0, "reprovadas_decadas": 0,
            "reprovadas_repeticao": 0, "reprovadas_cores": 0,
            "reprovadas_regra31": 0, "reprovadas_progressao": 0,
        }

    def verificar(self, chave: list) -> bool:
        self.stats["testadas"] += 1
        chave_s = sorted(chave)

        # ── Filter A – Configurable sum range ────────────────────────────────
        if "A" in self.active_filters:
            s = sum(chave_s)
            if not (self.soma_min <= s <= self.soma_max):
                self.stats["reprovadas_soma"] += 1
                return False

        # ── Filter B – Consecutive pairs / triples ────────────────────────────
        if "B" in self.active_filters:
            pares = 0
            for i in range(len(chave_s) - 1):
                if chave_s[i+1] == chave_s[i] + 1:
                    pares += 1
                    if i < len(chave_s) - 2 and chave_s[i+2] == chave_s[i] + 2:
                        self.stats["reprovadas_consecutivos"] += 1
                        return False
            if pares > 1:
                self.stats["reprovadas_consecutivos"] += 1
                return False

        # ── Filter C – Same final digit (max 2 share same last digit) ─────────
        if "C" in self.active_filters:
            finais = [n % 10 for n in chave_s]
            if max(Counter(finais).values()) > 2:
                self.stats["reprovadas_finais"] += 1
                return False

        # ── Filter D – Decades spread (min 3 different decades) ──────────────
        if "D" in self.active_filters:
            decadas = {(n - 1) // 10 for n in chave_s}
            if len(decadas) < 3:
                self.stats["reprovadas_decadas"] += 1
                return False

        # ── Filter E – Repeat from last draw (max 2 shared numbers) ──────────
        if "E" in self.active_filters and self.ultimo_sorteio:
            comuns = set(chave_s).intersection(set(self.ultimo_sorteio))
            if len(comuns) > 2:
                self.stats["reprovadas_repeticao"] += 1
                return False

        # ── Filter F – Colour system (last 9 draws) ───────────────────────────
        if "F" in self.active_filters:
            v = self.cores.get("vermelhos", set())
            g = self.cores.get("verdes", set())
            a = self.cores.get("azuis", set())
            c = self.cores.get("castanhos", set())

            qtd_v = sum(1 for n in chave_s if n in v)
            qtd_g = sum(1 for n in chave_s if n in g)
            qtd_a = sum(1 for n in chave_s if n in a)
            qtd_c = sum(1 for n in chave_s if n in c)

            if not (1 <= qtd_v <= 3): self.stats["reprovadas_cores"] += 1; return False
            if not (1 <= qtd_g <= 3): self.stats["reprovadas_cores"] += 1; return False
            if not (0 <= qtd_a <= 2): self.stats["reprovadas_cores"] += 1; return False
            if qtd_c > 0:             self.stats["reprovadas_cores"] += 1; return False

        # ── Filter G – Regra do 31 (anti-calendário) ─────────────────────────
        if "G" in self.active_filters:
            acima_31 = sum(1 for n in chave_s if n > 31)
            if acima_31 < 1:
                self.stats["reprovadas_regra31"] += 1
                return False

        # ── Filter H – Perfect arithmetic progression ─────────────────────────
        if "H" in self.active_filters:
            diffs = [chave_s[i+1] - chave_s[i] for i in range(4)]
            if len(set(diffs)) == 1:
                self.stats["reprovadas_progressao"] += 1
                return False

        self.stats["aprovadas"] += 1
        return True

    def reset_stats(self):
        for k in self.stats:
            self.stats[k] = 0

    def resumo_filtros(self) -> list[dict]:
        """Return metadata + stats for every filter (including inactive ones)."""
        result = []
        for f_id in self.ALL_FILTERS:
            info = self.FILTER_INFO[f_id]
            config_str = {
                "A": f"{self.soma_min}–{self.soma_max}",
                "B": "máx 1 par, 0 triplos",
                "C": "máx 2 iguais",
                "D": "mín 3 décadas",
                "E": "máx 2 do anterior",
                "F": "V:1-3 | G:1-3 | A:0-1 | C:0",
                "G": "mín 1 número > 31",
                "H": "rejeitar sequências perfeitas",
            }[f_id]
            result.append({
                "id":        f_id,
                "nome":      info["nome"],
                "descricao": info["descricao"],
                "config":    config_str,
                "reprovadas": self.stats[self.STAT_MAP[f_id]],
                "ativo":     f_id in self.active_filters,
            })
        return result


# ═════════════════════════════════════════════════════════════════════════════
# KEY GENERATOR
# ═════════════════════════════════════════════════════════════════════════════
class KeyGenerator:
    def __init__(self, filter_engine: FilterEngine, db: DatabaseManager):
        self.fe = filter_engine
        self.db = db
        self.contagem_estrelas = {i: 0 for i in range(1, 13)}
        self._load_star_counts()

    def _load_star_counts(self):
        raw = self.db.get_metadata("contagem_estrelas")
        if raw:
            try:
                loaded = json.loads(raw)
                for k, v in loaded.items():
                    self.contagem_estrelas[int(k)] = v
            except (json.JSONDecodeError, ValueError, KeyError):
                pass

    def _save_star_counts(self):
        self.db.set_metadata("contagem_estrelas", json.dumps(self.contagem_estrelas))

    def escolher_estrelas_equilibradas(self) -> list:
        disponiveis = list(range(1, 13))
        random.shuffle(disponiveis)
        disponiveis.sort(key=lambda x: self.contagem_estrelas[x])
        e1, e2 = disponiveis[0], disponiveis[1]
        self.contagem_estrelas[e1] += 1
        self.contagem_estrelas[e2] += 1
        return sorted([e1, e2])

    def gerar_chave(self, max_tentativas: int = 100000) -> dict | None:
        self.fe.reset_stats()
        padrao = random.choice(PADROES_EQUILIBRADOS)
        bi_n, bp_n, ai_n, ap_n = padrao

        # Safety net: rotate pattern every 2000 attempts to avoid rare
        # dead-end combinations where a pattern + active filters conflict.
        for i in range(max_tentativas):
            if i > 0 and i % 2000 == 0:
                padrao = random.choice(PADROES_EQUILIBRADOS)
                bi_n, bp_n, ai_n, ap_n = padrao
            try:
                nums = (
                    random.sample(BI, bi_n) +
                    random.sample(BP, bp_n) +
                    random.sample(AI, ai_n) +
                    random.sample(AP, ap_n)
                )
            except ValueError:
                continue

            if len(set(nums)) != 5:
                continue

            nums_sorted = sorted(nums)
            if self.fe.verificar(nums_sorted):
                estrelas = self.escolher_estrelas_equilibradas()
                self._save_star_counts()
                self.db.guardar_chave_gerada(nums_sorted, estrelas, padrao)
                return {
                    "numeros": nums_sorted,
                    "estrelas": estrelas,
                    "padrao": padrao,
                    "soma": sum(nums_sorted),
                    "tentativas": self.fe.stats["testadas"],
                }

        return None

    def gerar_multiplas_chaves(self, quantidade: int) -> list[dict]:
        chaves = []
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[cyan]{task.completed}/{task.total}"),
            console=console,
        ) as progress:
            task = progress.add_task("[yellow]A gerar chaves...", total=quantidade)
            tentativas_total = 0
            for i in range(quantidade):
                chave = self.gerar_chave()
                if chave:
                    chaves.append(chave)
                    tentativas_total += chave["tentativas"]
                progress.advance(task)
        return chaves


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ═════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    # Colours matching the system
    FILL_VERMELHO  = PatternFill("solid", fgColor="FF9999")
    FILL_VERDE     = PatternFill("solid", fgColor="99FF99")
    FILL_AZUL      = PatternFill("solid", fgColor="99CCFF")
    FILL_CASTANHO  = PatternFill("solid", fgColor="D2A679")
    FILL_HEADER    = PatternFill("solid", fgColor="2E4057")
    FILL_STAR      = PatternFill("solid", fgColor="FFD700")
    FILL_SUM       = PatternFill("solid", fgColor="E8E8E8")
    FILL_ROW_ALT   = PatternFill("solid", fgColor="F5F5F5")

    def exportar(self, chaves: list[dict], cores: dict,
                 nome_ficheiro: str | None = None,
                 filtros: list[dict] | None = None,
                 config: dict | None = None) -> Path:
        if not nome_ficheiro:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_ficheiro = f"chaves_{ts}.xlsx"

        filepath = EXCEL_DIR / nome_ficheiro
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chaves Geradas"

        # ── Title block ──────────────────────────────────────────────────────
        ws.merge_cells("C3:J3")
        t = ws["C3"]
        t.value = "EUROMILHÕES – CHAVES GERADAS"
        t.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.fill = PatternFill("solid", fgColor="2E4057")
        ws.row_dimensions[3].height = 30

        ws.merge_cells("C4:J4")
        sub = ws["C4"]
        sub.value = f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(chaves)} chaves"
        sub.font = Font(italic=True, size=10, color="666666")
        sub.alignment = Alignment(horizontal="center")

        # ── Legend ───────────────────────────────────────────────────────────
        ws.merge_cells("C6:J6")
        ws["C6"].value = "LEGENDA DE CORES"
        ws["C6"].font = Font(bold=True, size=10, color="2E4057")

        legend_items = [
            ("C7", "VERMELHOS (0×)", "FF9999"),
            ("E7", "VERDES (1×)", "99FF99"),
            ("G7", "AZUIS (2×)", "99CCFF"),
            ("I7", "CASTANHOS (3×+) – EXCLUÍDOS", "D2A679"),
        ]
        for cell_ref, label, colour in legend_items:
            cell = ws[cell_ref]
            cell.value = label
            cell.fill = PatternFill("solid", fgColor=colour)
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal="center")

        # ── Column headers at row 11 ─────────────────────────────────────────
        headers = ["N1", "N2", "N3", "N4", "N5", "E1", "E2", "SOMA"]
        thin = Side(style="thin", color="AAAAAA")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for col_idx, header in enumerate(headers, start=3):  # C=3
            cell = ws.cell(row=11, column=col_idx, value=header)
            cell.fill = self.FILL_HEADER
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = 8

        ws.row_dimensions[11].height = 22

        # ── Data rows starting at row 12 ─────────────────────────────────────
        for row_num, chave in enumerate(chaves, start=12):
            fill_row = self.FILL_ROW_ALT if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            numeros = chave["numeros"]
            estrelas = chave["estrelas"]

            for col_offset, num in enumerate(numeros):  # cols C–G (3–7)
                cell = ws.cell(row=row_num, column=3 + col_offset, value=num)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                # Colour by classification
                if num in cores.get("vermelhos", set()):
                    cell.fill = self.FILL_VERMELHO
                elif num in cores.get("verdes", set()):
                    cell.fill = self.FILL_VERDE
                elif num in cores.get("azuis", set()):
                    cell.fill = self.FILL_AZUL
                elif num in cores.get("castanhos", set()):
                    cell.fill = self.FILL_CASTANHO
                else:
                    cell.fill = fill_row

            for col_offset, est in enumerate(estrelas):  # cols H–I (8–9)
                cell = ws.cell(row=row_num, column=8 + col_offset, value=est)
                cell.fill = self.FILL_STAR
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            soma_cell = ws.cell(row=row_num, column=10, value=chave["soma"])
            soma_cell.fill = self.FILL_SUM
            soma_cell.alignment = Alignment(horizontal="center")
            soma_cell.border = border
            soma_cell.font = Font(bold=True, color="2E4057")

        # ── Freeze panes ─────────────────────────────────────────────────────
        ws.freeze_panes = "C12"

        # ── Stats sheet ──────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Estatísticas")
        ws2["A1"] = "DISTRIBUIÇÃO POR CORES"
        ws2["A1"].font = Font(bold=True, size=12)

        all_nums = [n for ch in chaves for n in ch["numeros"]]
        colour_counts = {
            "VERMELHOS": sum(1 for n in all_nums if n in cores.get("vermelhos", set())),
            "VERDES":    sum(1 for n in all_nums if n in cores.get("verdes", set())),
            "AZUIS":     sum(1 for n in all_nums if n in cores.get("azuis", set())),
        }
        for r, (label, count) in enumerate(colour_counts.items(), start=2):
            ws2.cell(row=r, column=1, value=label)
            ws2.cell(row=r, column=2, value=count)

        somas = [ch["soma"] for ch in chaves]
        ws2["A7"] = "SOMAS"
        ws2["A7"].font = Font(bold=True)
        ws2["A8"]  = "Mínima";  ws2["B8"]  = min(somas)
        ws2["A9"]  = "Máxima";  ws2["B9"]  = max(somas)
        ws2["A10"] = "Média";   ws2["B10"] = round(sum(somas)/len(somas), 1)

        # ── Filters used ─────────────────────────────────────────────────────
        if filtros:
            ws2["A12"] = "FILTROS UTILIZADOS"
            ws2["A12"].font = Font(bold=True, size=11)
            ws2["B12"] = "ESTADO"
            ws2["B12"].font = Font(bold=True, size=11)
            ws2["C12"] = "CONFIGURAÇÃO"
            ws2["C12"].font = Font(bold=True, size=11)
            for r, f in enumerate(filtros, start=13):
                ws2.cell(row=r, column=1, value=f"[{f['id']}] {f['nome']}")
                status = "ACTIVO" if f.get("ativo", True) else "DESLIGADO"
                cell_s = ws2.cell(row=r, column=2, value=status)
                cell_s.font = Font(
                    color="00AA00" if f.get("ativo", True) else "AA3333",
                    bold=True
                )
                ws2.cell(row=r, column=3, value=f.get("config", ""))
            ws2.column_dimensions["A"].width = 28
            ws2.column_dimensions["B"].width = 12
            ws2.column_dimensions["C"].width = 32

        wb.save(filepath)
        return filepath


# ═════════════════════════════════════════════════════════════════════════════
# TERMINAL UI
# ═════════════════════════════════════════════════════════════════════════════
class TerminalUI:
    def __init__(self):
        self.db = DatabaseManager()
        self.scraper = EuromilhoesScraper()
        self.stats = StatisticsAnalyzer(self.db)
        self.exporter = ExcelExporter()

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _header(self):
        console.print()
        console.print(Panel(
            Text.from_markup(
                f"[bold cyan]EUROMILHÕES[/] [dim]v{VERSION}[/]  |  "
                f"[yellow]Gerador Profissional de Chaves[/]\n"
                f"[dim]Baseado na metodologia Lotterycodex BI-BP-AI-AP[/]"
            ),
            style="bold blue", box=box.DOUBLE_EDGE, padding=(0, 4)
        ))

    def _status_bar(self):
        total = self.db.total_sorteios()
        ultimo = self.db.ultimo_sorteio()
        data_ult = ultimo["data"] if ultimo else "N/D"
        console.print(
            f"[dim]Base de dados:[/] [green]{total}[/] sorteios  |  "
            f"[dim]Último sorteio:[/] [cyan]{data_ult}[/]",
            justify="center"
        )
        console.print()

    def _cor_tag(self, num: int, cores: dict) -> str:
        if num in cores.get("vermelhos", set()): return f"[bold red]{num:2d}[/]"
        if num in cores.get("verdes", set()):    return f"[bold green]{num:2d}[/]"
        if num in cores.get("azuis", set()):     return f"[bold blue]{num:2d}[/]"
        if num in cores.get("castanhos", set()): return f"[bold #8B4513]{num:2d}[/]"
        return f"[white]{num:2d}[/]"

    # ── Main menu ─────────────────────────────────────────────────────────────
    def menu_principal(self):
        while True:
            self._header()
            self._status_bar()

            table = Table(box=box.ROUNDED, show_header=False, padding=(0,2))
            table.add_column(style="bold cyan", width=4)
            table.add_column()
            items = [
                ("1", "Gerar chaves"),
                ("2", "Ver último sorteio + classificação de cores"),
                ("3", "Actualizar base de dados (web scraping)"),
                ("4", "Inserir sorteio manualmente"),
                ("5", "Estatísticas e análise"),
                ("6", "Ver estratégias utilizadas"),
                ("7", "Importar histórico via CSV"),
                ("8", "Backtesting (testar filtros no histórico)"),
                ("0", "[dim]Sair[/]"),
            ]
            for num, label in items:
                table.add_row(f"[{num}]", label)
            console.print(table, justify="center")
            console.print()

            choice = Prompt.ask("[bold]Opção", default="1")
            if   choice == "1": self.menu_gerar()
            elif choice == "2": self.ver_ultimo_sorteio()
            elif choice == "3": self.actualizar_db_web()
            elif choice == "4": self.inserir_manual()
            elif choice == "5": self.menu_estatisticas()
            elif choice == "6": self.ver_estrategias()
            elif choice == "7": self.importar_csv()
            elif choice == "8": self.backtesting()
            elif choice == "0": break
            else:
                console.print("[red]Opção inválida.[/]")

    # ── Generate keys ─────────────────────────────────────────────────────────
    def menu_gerar(self):
        console.clear()
        self._header()

        # Prepare colours
        ultimos_9 = self.db.ultimos_n_sorteios(COLOR_WINDOW_DRAWS)
        if len(ultimos_9) < 9:
            console.print(f"[yellow]Aviso:[/] Apenas {len(ultimos_9)} sorteios na BD (recomendado: 9+).")

        cores = self.stats.classificar_cores(ultimos_9)
        ultimo = self.db.ultimo_sorteio()
        ultimo_nums = ultimo["numeros"] if ultimo else []

        # Show colour classification
        self._mostrar_cores(cores)

        qtd = IntPrompt.ask("[bold]Quantas chaves gerar?", default=10)
        if qtd < 1 or qtd > 200:
            console.print("[red]Valor inválido (1–200).[/]")
            return

        fe = FilterEngine(cores, ultimo_nums)
        gen = KeyGenerator(fe, self.db)

        chaves = gen.gerar_multiplas_chaves(qtd)

        if not chaves:
            console.print("[red]Não foi possível gerar chaves com os filtros actuais.[/]")
            return

        console.print()
        self._mostrar_chaves(chaves, cores)

        # Export to Excel?
        if Confirm.ask("[bold]Exportar para Excel?", default=True):
            path = self.exporter.exportar(chaves, cores)
            console.print(f"[green]Ficheiro guardado:[/] {path}")

        input("\nPressiona Enter para continuar...")

    def _mostrar_cores(self, cores: dict):
        table = Table(title="Classificação de Cores (últimos 9 sorteios)",
                      box=box.SIMPLE_HEAD)
        table.add_column("Cor", style="bold")
        table.add_column("Critério")
        table.add_column("Números", overflow="fold")
        table.add_column("Qtd", justify="right")
        table.add_column("Regra combo")

        rows = [
            ("VERMELHOS", "0 aparições", sorted(cores["vermelhos"]), "bold red", "1–3"),
            ("VERDES",    "1 aparição",  sorted(cores["verdes"]),    "bold green", "1–3"),
            ("AZUIS",     "2 aparições", sorted(cores["azuis"]),     "bold blue",  "0–2"),
            ("CASTANHOS", "3+ aparições",sorted(cores["castanhos"]), "bold #8B4513", "EXCLUÍDOS"),
        ]
        for nome, crit, nums, style, regra in rows:
            nums_str = " ".join(str(n) for n in nums)
            table.add_row(
                f"[{style}]{nome}[/]", crit, f"[dim]{nums_str}[/]",
                str(len(nums)), regra
            )
        console.print(table)
        console.print()

    def _mostrar_chaves(self, chaves: list[dict], cores: dict):
        table = Table(
            title=f"[bold cyan]{len(chaves)} Chaves Geradas[/]",
            box=box.ROUNDED, show_lines=False
        )
        for col in ["#", "N1", "N2", "N3", "N4", "N5", "E1", "E2", "Soma", "Padrão [BI BP AI AP]"]:
            table.add_column(col, justify="center")

        for i, ch in enumerate(chaves, 1):
            nums_tagged = [self._cor_tag(n, cores) for n in ch["numeros"]]
            stars = [f"[bold yellow]{e}[/]" for e in ch["estrelas"]]
            padrao_str = str(ch["padrao"]).replace("[","").replace("]","")
            table.add_row(
                str(i), *nums_tagged, *stars,
                f"[bold]{ch['soma']}[/]", padrao_str
            )
        console.print(table)

    # ── Last draw ─────────────────────────────────────────────────────────────
    def ver_ultimo_sorteio(self):
        console.clear()
        self._header()
        ultimo = self.db.ultimo_sorteio()
        if not ultimo:
            console.print("[yellow]Nenhum sorteio na base de dados.[/]")
        else:
            cores = self.stats.classificar_cores(self.db.ultimos_n_sorteios(COLOR_WINDOW_DRAWS))
            nums_str = "  ".join(self._cor_tag(n, cores) for n in ultimo["numeros"])
            stars_str = "  ".join(f"[bold yellow]{e}[/]" for e in ultimo["estrelas"])
            console.print(Panel(
                f"[dim]Data:[/] [cyan]{ultimo['data']}[/]\n\n"
                f"Números: {nums_str}\n"
                f"Estrelas: {stars_str}\n"
                f"Soma: [bold]{ultimo['soma']}[/]",
                title="Último Sorteio", border_style="cyan"
            ))
            console.print()
            self._mostrar_cores(cores)

        console.print()
        input("Pressiona Enter para continuar...")

    # ── Web update ────────────────────────────────────────────────────────────
    def actualizar_db_web(self):
        console.clear()
        self._header()
        console.print("[cyan]A tentar obter o último sorteio via internet...[/]\n")
        with console.status("[bold green]A contactar servidores..."):
            result = self.scraper.fetch_ultimo_sorteio()

        if result:
            console.print(Panel(
                f"Data: [cyan]{result['data']}[/]\n"
                f"Números: [bold]{result['numeros']}[/]\n"
                f"Estrelas: [bold yellow]{result['estrelas']}[/]",
                title="Sorteio encontrado", border_style="green"
            ))
            if Confirm.ask("Guardar na base de dados?", default=True):
                ok = self.db.inserir_sorteio(result["data"], result["numeros"],
                                              result["estrelas"], fonte="web")
                console.print("[green]Guardado![/]" if ok else "[yellow]Já existe ou erro.[/]")
        else:
            console.print(
                "[red]Não foi possível obter dados automaticamente.[/]\n"
                "[dim]Os sites de lotaria podem ter alterado a sua estrutura.\n"
                "Usa a opção 4 para inserir manualmente.[/]"
            )

        input("\nPressiona Enter para continuar...")

    # ── Manual insert ─────────────────────────────────────────────────────────
    def inserir_manual(self):
        console.clear()
        self._header()
        console.print("[bold]Inserir sorteio manualmente[/]\n")

        data = Prompt.ask("Data (YYYY-MM-DD)", default=datetime.date.today().isoformat())
        try:
            datetime.date.fromisoformat(data)
        except ValueError:
            console.print("[red]Data inválida.[/]"); return

        console.print("Introduz os 5 números principais (1–50), separados por vírgulas:")
        try:
            nums_raw = Prompt.ask("Números")
            nums = sorted([int(x.strip()) for x in nums_raw.split(",")])
            if len(nums) != 5 or len(set(nums)) != 5:
                raise ValueError
            if not all(1 <= n <= 50 for n in nums):
                raise ValueError
        except (ValueError, TypeError):
            console.print("[red]Números inválidos. Precisas de 5 números únicos entre 1 e 50.[/]")
            return

        console.print("Introduz as 2 estrelas (1–12), separadas por vírgulas:")
        try:
            stars_raw = Prompt.ask("Estrelas")
            stars = sorted([int(x.strip()) for x in stars_raw.split(",")])
            if len(stars) != 2 or len(set(stars)) != 2:
                raise ValueError
            if not all(1 <= s <= 12 for s in stars):
                raise ValueError
        except (ValueError, TypeError):
            console.print("[red]Estrelas inválidas. Precisas de 2 estrelas únicas entre 1 e 12.[/]")
            return

        ok = self.db.inserir_sorteio(data, nums, stars, fonte="manual")
        console.print("[green]Sorteio guardado![/]" if ok else "[yellow]Já existe na BD.[/]")
        input("\nPressiona Enter para continuar...")

    # ── Statistics ─────────────────────────────────────────────────────────────
    def menu_estatisticas(self):
        while True:
            console.clear()
            self._header()
            console.print("[bold]ESTATÍSTICAS[/]\n")
            choice = Prompt.ask(
                "1) Frequência números  2) Frequência estrelas  3) Análise de padrões  "
                "4) Números atrasados  5) Sequências quentes  0) Voltar",
                default="0"
            )
            if   choice == "1": self._stats_freq_numeros()
            elif choice == "2": self._stats_freq_estrelas()
            elif choice == "3": self._stats_padroes()
            elif choice == "4": self._stats_atrasados()
            elif choice == "5": self._stats_quentes()
            elif choice == "0": break

    def _stats_freq_numeros(self):
        console.clear()
        self._header()
        freq = self.stats.frequencia_numeros()
        total = self.db.total_sorteios()
        if not freq:
            console.print("[yellow]Sem dados.[/]"); input(); return

        table = Table(title="Frequência dos Números (1–50)", box=box.SIMPLE_HEAD)
        table.add_column("Número", justify="right")
        table.add_column("Vezes", justify="right")
        table.add_column("% dos sorteios", justify="right")
        table.add_column("Barra")

        for n in range(1, 51):
            v = freq.get(n, 0)
            pct = (v / total * 100) if total > 0 else 0
            bar_len = int(pct / 2)
            table.add_row(str(n), str(v), f"{pct:.1f}%", "█" * bar_len)
        console.print(table)
        input("\nPressiona Enter...")

    def _stats_freq_estrelas(self):
        console.clear()
        self._header()
        freq = self.stats.frequencia_estrelas()
        total = self.db.total_sorteios()
        table = Table(title="Frequência das Estrelas (1–12)", box=box.SIMPLE_HEAD)
        table.add_column("Estrela", justify="right")
        table.add_column("Vezes", justify="right")
        table.add_column("% dos sorteios", justify="right")
        for n in range(1, 13):
            v = freq.get(n, 0)
            pct = (v / total * 100) if total > 0 else 0
            table.add_row(str(n), str(v), f"{pct:.1f}%")
        console.print(table)
        input("\nPressiona Enter...")

    def _stats_padroes(self):
        console.clear()
        self._header()
        padroes = self.stats.analise_padroes()
        total = self.db.total_sorteios()
        table = Table(title="Padrões BI-BP-AI-AP Históricos", box=box.SIMPLE_HEAD)
        table.add_column("Padrão [BI,BP,AI,AP]")
        table.add_column("Equilibrado?", justify="center")
        table.add_column("Ocorrências", justify="right")
        table.add_column("% histórico", justify="right")

        equilibrados_set = {tuple(p) for p in PADROES_EQUILIBRADOS}
        for padrao, count in sorted(padroes.items(), key=lambda x: -x[1]):
            eq = "[green]✓[/]" if padrao in equilibrados_set else "[red]✗[/]"
            pct = (count / total * 100) if total > 0 else 0
            table.add_row(str(list(padrao)), eq, str(count), f"{pct:.1f}%")
        console.print(table)
        input("\nPressiona Enter...")

    def _stats_atrasados(self):
        console.clear()
        self._header()
        atrasados = self.stats.numeros_atrasados(15)
        table = Table(title="Números Mais Atrasados", box=box.SIMPLE_HEAD)
        table.add_column("Número", justify="right")
        table.add_column("Há quantos sorteios?", justify="right")
        for num, draws_ago in atrasados:
            table.add_row(str(num), str(draws_ago))
        console.print(table)
        input("\nPressiona Enter...")

    def _stats_quentes(self):
        console.clear()
        self._header()
        quentes = self.stats.sequencias_quentes(10)
        table = Table(title="Números Mais Frequentes (últimos 10 sorteios)", box=box.SIMPLE_HEAD)
        table.add_column("Número", justify="right")
        table.add_column("Aparições", justify="right")
        for num, count in quentes:
            table.add_row(str(num), str(count))
        console.print(table)
        input("\nPressiona Enter...")

    # ── Strategies ────────────────────────────────────────────────────────────
    def ver_estrategias(self):
        console.clear()
        self._header()
        console.print(Panel(
            """[bold cyan]ESTRATÉGIAS E FILTROS ACTIVOS[/]

[bold yellow]METODOLOGIA BASE: Lotterycodex BI-BP-AI-AP[/]
Os 50 números são divididos em 4 quadrantes:
  [red]BI[/] (Baixos-Ímpares) : { """ + ", ".join(str(n) for n in BI) + """ }
  [green]BP[/] (Baixos-Pares)   : { """ + ", ".join(str(n) for n in BP) + """ }
  [blue]AI[/] (Altos-Ímpares)  : { """ + ", ".join(str(n) for n in AI) + """ }
  [cyan]AP[/] (Altos-Pares)    : { """ + ", ".join(str(n) for n in AP) + """ }

[bold yellow]16 PADRÕES EQUILIBRADOS[/]
Apenas combinações com distribuição equilibrada entre quadrantes.
Estes 16 padrões cobrem ~1.333.800 combinações possíveis.

[bold yellow]FILTROS ACTIVOS:[/]
  [A] Soma configurável: padrão 80–190 | apertado 95–160
      → ~93% dos sorteios históricos têm soma neste intervalo
  [B] Máx 1 par consecutivo, 0 triplos consecutivos
      → pares: ~42% dos sorteios | triplos: <1% → sempre eliminar
  [C] Máx 2 números com o mesmo dígito final
      → 3+ iguais representam <4% dos sorteios históricos
  [D] Mín 3 dezenas diferentes representadas (1-10, 11-20, 21-30, 31-40, 41-50)
      → evita concentração de números numa zona do boletim
  [E] Máx 2 repetições do sorteio anterior
      → restrição de curto prazo baseada em tendências
  [F] Sistema de cores (últimos 9 sorteios):
      • VERMELHOS (0×) → 1 a 3 na chave
      • VERDES    (1×) → 1 a 3 na chave
      • AZUIS     (2×) → 0 a 2 na chave
      • CASTANHOS (3×+) → EXCLUÍDOS (0 na chave)
  [G] Regra do 31 – mínimo 1 número acima de 31
      → a maioria dos jogadores usa aniversários (1–31);
        forçar números altos reduz partilha do jackpot
  [H] Rejeitar progressões aritméticas perfeitas
      → ex: 5,10,15,20,25 ou 3,9,15,21,27 são escolhas
        demasiado populares e devem ser evitadas

[bold yellow]ESTRELAS EQUILIBRADAS[/]
O sistema memoriza quantas vezes cada estrela (1–12) foi usada
e selecciona sempre as 2 com menor contagem → distribuição uniforme.
""",
            title="Estratégias", border_style="cyan", padding=(1, 2)
        ))
        input("Pressiona Enter para continuar...")

    # ── CSV import ────────────────────────────────────────────────────────────
    def importar_csv(self):
        console.clear()
        self._header()
        console.print(
            "[bold]Importar histórico via CSV[/]\n"
            "[dim]Formato esperado: data,n1,n2,n3,n4,n5,e1,e2\n"
            "Exemplo: 2024-01-05,3,12,24,37,45,2,9[/]\n"
        )
        path_str = Prompt.ask("Caminho do ficheiro CSV")
        path = Path(path_str)
        if not path.exists():
            console.print("[red]Ficheiro não encontrado.[/]")
            input(); return

        imported = 0
        errors = 0
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) < 8:
                    errors += 1; continue
                try:
                    data = row[0].strip()
                    nums = [int(row[i]) for i in range(1, 6)]
                    stars = [int(row[i]) for i in range(6, 8)]
                    if self.db.inserir_sorteio(data, nums, stars, fonte="csv"):
                        imported += 1
                    else:
                        errors += 1
                except (ValueError, IndexError):
                    errors += 1

        console.print(f"[green]Importados:[/] {imported}  |  [red]Erros:[/] {errors}")
        input("\nPressiona Enter para continuar...")

    # ── Backtesting ───────────────────────────────────────────────────────────
    def backtesting(self):
        console.clear()
        self._header()
        console.print(
            "[bold]BACKTESTING[/]\n"
            "[dim]Verifica quantos sorteios históricos passariam nos filtros actuais.[/]\n"
        )
        todos = self.db.todos_sorteios()
        if len(todos) < 10:
            console.print("[yellow]Precisas de pelo menos 10 sorteios para backtesting.[/]")
            input(); return

        passaram = 0
        falharam = 0
        reprovados = Counter()

        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TextColumn("{task.completed}/{task.total}"),
                      console=console) as progress:
            task = progress.add_task("[cyan]A analisar...", total=len(todos))

            for i, sorteio in enumerate(todos):
                # Use the 9 draws BEFORE this one for colour classification
                contexto = todos[i+1:i+10] if i + 10 <= len(todos) else todos[i+1:]
                cores = self.stats.classificar_cores(contexto)
                anterior = todos[i+1]["numeros"] if i + 1 < len(todos) else []
                fe = FilterEngine(cores, anterior)
                fe.reset_stats()

                if fe.verificar(sorteio["numeros"]):
                    passaram += 1
                else:
                    for key, val in fe.stats.items():
                        if key.startswith("reprovadas_") and val > 0:
                            reprovados[key] += 1

                falharam = (i + 1) - passaram
                progress.advance(task)

        total = len(todos)
        pct = (passaram / total * 100) if total > 0 else 0

        console.print(Panel(
            f"Total sorteios analisados: [bold]{total}[/]\n"
            f"Passaram nos filtros: [bold green]{passaram}[/] ({pct:.1f}%)\n"
            f"Reprovaram: [bold red]{falharam}[/]\n\n"
            + "\n".join(f"  {k.replace('reprovadas_','')}: {v}" for k, v in reprovados.most_common()),
            title="Resultado do Backtesting", border_style="cyan"
        ))
        input("\nPressiona Enter para continuar...")


# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════
def main():
    console.clear()
    ui = TerminalUI()
    try:
        ui.menu_principal()
    except KeyboardInterrupt:
        console.print("\n[dim]Saindo...[/]")
    console.print("\n[bold cyan]Obrigado por usar o Gerador de Chaves EuroMillhões v8![/]\n")


if __name__ == "__main__":
    main()
